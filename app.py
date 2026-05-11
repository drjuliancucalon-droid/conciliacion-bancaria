"""
CREDIEXPRESS POPAYÁN SAS — Conciliación Bancaria Interactiva Premium
Soporte multiformato (PDF, CSV, Excel, TXT) + OCR para PDF escaneados
100% fiel al notebook original en procesamiento y reglas de negocio
"""

import streamlit as st
import sqlite3, json
from datetime import datetime
import re, io, warnings, os, tempfile, logging
from pathlib import Path
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import pdfplumber
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# OCR (opcional, solo se usa si las librerías están instaladas)
try:
    from pdf2image import convert_from_path
    import pytesseract
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False

warnings.filterwarnings('ignore')
pd.set_option('display.float_format', lambda x: f'{x:,.2f}')
pd.set_option('display.max_colwidth', 90)
pd.set_option('display.max_rows', 800)

st.set_page_config(page_title="Conciliación CREDIEXPRESS", page_icon="🏦", layout="wide")
# ══════════════════════════════════════════════════════════════════════════════
# MÓDULO DE ALMACENAMIENTO — SQLite (offline) / Google Sheets (cloud)
# ══════════════════════════════════════════════════════════════════════════════

BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
OFFLINE_MODE = not os.path.exists("/mount/src")   # True=local, False=Streamlit Cloud
DB_PATH      = os.path.join(BASE_DIR, "conciliaciones.db")

# ── SQLite ────────────────────────────────────────────────────────────────────
def _init_db():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""CREATE TABLE IF NOT EXISTS historial (
        id               INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha_hora       TEXT,
        archivo_banco    TEXT,
        archivo_auxiliar TEXT,
        periodo          TEXT,
        n_banco          INTEGER,
        n_aux            INTEGER,
        n_exactas        INTEGER,
        n_aprox          INTEGER,
        n_solo_banco     INTEGER,
        n_solo_aux       INTEGER,
        tasa             REAL,
        saldo_banco      REAL,
        saldo_aux        REAL,
        diferencia_neta  REAL,
        excel_path       TEXT
    )""")
    # ── Fase C: catálogo de formatos PDF aprendidos ────────────────────────
    conn.execute("""CREATE TABLE IF NOT EXISTS pdf_formatos (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        firma           TEXT UNIQUE,
        tipo_doc        TEXT,
        columnas        TEXT,
        fmt_fecha       TEXT,
        prefijos_doc    TEXT,
        banco_detectado TEXT,
        usos            INTEGER DEFAULT 1,
        ultima_vez      TEXT
    )""")
    # ── Fase D: catalogo de conceptos NC aprendidos ──────────────────────
    conn.execute("""CREATE TABLE IF NOT EXISTS nc_catalogo (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        uuid            TEXT UNIQUE,
        banco_tokens    TEXT,
        aux_tokens      TEXT,
        confirmaciones  INTEGER DEFAULT 1,
        nivel           TEXT DEFAULT 'PENDIENTE',
        aprobado_por    TEXT DEFAULT 'AUTO',
        fecha_primera   TEXT,
        fecha_ultima    TEXT,
        sync_status     TEXT DEFAULT 'PENDIENTE_SYNC'
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS nc_aprendizaje (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        uuid            TEXT UNIQUE,
        banco_desc_raw  TEXT,
        aux_concepto_raw TEXT,
        banco_tokens    TEXT,
        aux_tokens      TEXT,
        veces_visto     INTEGER DEFAULT 1,
        fecha_primera   TEXT,
        fecha_ultima    TEXT
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS nc_historial_match (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha           TEXT,
        banco_desc      TEXT,
        aux_doc         TEXT,
        aux_concepto    TEXT,
        metodo          TEXT,
        valor_banco     REAL,
        valor_aux       REAL
    )""")
    conn.commit()
    return conn

def _auto_guardar_archivo(uploaded_file, subfolder="datos_entrada"):
    if not OFFLINE_MODE or uploaded_file is None:
        return None, False
    dest_dir = os.path.join(BASE_DIR, subfolder, datetime.now().strftime("%Y-%m"))
    os.makedirs(dest_dir, exist_ok=True)
    dest = os.path.join(dest_dir, uploaded_file.name)
    if not os.path.exists(dest):
        with open(dest, "wb") as f:
            f.write(uploaded_file.getvalue())
        return dest, True
    return dest, False

def _auto_guardar_excel(excel_bytes, nombre):
    if not OFFLINE_MODE:
        return None
    ts      = datetime.now().strftime("%Y-%m-%d_%H-%M")
    destdir = os.path.join(BASE_DIR, "reportes_excel", ts)
    os.makedirs(destdir, exist_ok=True)
    dest = os.path.join(destdir, nombre)
    with open(dest, "wb") as f:
        f.write(excel_bytes)
    return dest

def _guardar_historial_sqlite(d):
    try:
        conn = _init_db()
        conn.execute("""INSERT INTO historial
            (fecha_hora,archivo_banco,archivo_auxiliar,periodo,
             n_banco,n_aux,n_exactas,n_aprox,n_solo_banco,n_solo_aux,
             tasa,saldo_banco,saldo_aux,diferencia_neta,excel_path)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (d["fecha_hora"],d["archivo_banco"],d["archivo_auxiliar"],d["periodo"],
             d["n_banco"],d["n_aux"],d["n_exactas"],d["n_aprox"],
             d["n_solo_banco"],d["n_solo_aux"],d["tasa"],
             d["saldo_banco"],d["saldo_aux"],d["diferencia_neta"],d.get("excel_path","")))
        conn.commit(); conn.close()
    except Exception:
        pass

def _guardar_historial_sheets(d):
    try:
        import gspread
        from google.oauth2.service_account import Credentials
        creds_json = st.secrets.get("GOOGLE_SHEETS_CREDS", None)
        sheet_id   = st.secrets.get("GOOGLE_SHEET_ID",    None)
        if not creds_json or not sheet_id:
            return
        creds = Credentials.from_service_account_info(
            json.loads(creds_json),
            scopes=["https://spreadsheets.google.com/feeds",
                    "https://www.googleapis.com/auth/drive"])
        ws = gspread.authorize(creds).open_by_key(sheet_id).sheet1
        if not ws.get_all_values():
            ws.append_row(["Fecha","Banco","Auxiliar","Periodo",
                           "Mov.Banco","Asientos","Exactas","Aprox",
                           "Solo Banco","Solo Aux","Tasa %",
                           "Saldo Banco","Saldo Aux","Diferencia Neta"])
        ws.append_row([
            d["fecha_hora"],d["archivo_banco"],d["archivo_auxiliar"],d["periodo"],
            d["n_banco"],d["n_aux"],d["n_exactas"],d["n_aprox"],
            d["n_solo_banco"],d["n_solo_aux"],round(d["tasa"],1),
            round(d["saldo_banco"] or 0,2),round(d["saldo_aux"] or 0,2),
            round(d["diferencia_neta"] or 0,2)])
    except Exception:
        pass

def guardar_historial(d):
    """Punto de entrada: SQLite si es offline, Google Sheets si es cloud."""
    if OFFLINE_MODE:
        _guardar_historial_sqlite(d)
    else:
        _guardar_historial_sheets(d)

def leer_historial_sqlite(limite=8):
    try:
        if not os.path.exists(DB_PATH): return []
        conn = sqlite3.connect(DB_PATH)
        rows = conn.execute(
            """SELECT fecha_hora,archivo_banco,archivo_auxiliar,periodo,
                      tasa,n_exactas,n_banco,diferencia_neta
               FROM historial ORDER BY id DESC LIMIT ?""", (limite,)).fetchall()
        conn.close(); return rows
    except Exception: return []

# ── Fase C: Catálogo de formatos PDF aprendidos ───────────────────────────────
def _firma_pdf(nombre_archivo, n_columnas):
    """Genera una firma única por nombre normalizado + nro. columnas."""
    base = re.sub(r'[0-9_\-]', '', os.path.splitext(nombre_archivo or '')[0].lower()).strip()
    return f"{base}_{n_columnas}"

def registrar_formato_pdf(nombre_archivo, tipo_doc, columnas, fmt_fecha,
                          prefijos_doc, banco_detectado=""):
    """Guarda o actualiza el patrón de un PDF procesado exitosamente."""
    if not OFFLINE_MODE:
        return
    try:
        firma = _firma_pdf(nombre_archivo, len(columnas) if columnas else 0)
        ahora = datetime.now().isoformat(timespec='seconds')
        conn  = _init_db()
        existe = conn.execute(
            "SELECT id, usos FROM pdf_formatos WHERE firma=?", (firma,)).fetchone()
        if existe:
            conn.execute(
                "UPDATE pdf_formatos SET usos=?, ultima_vez=? WHERE firma=?",
                (existe[1] + 1, ahora, firma))
        else:
            conn.execute("""INSERT INTO pdf_formatos
                (firma, tipo_doc, columnas, fmt_fecha, prefijos_doc,
                 banco_detectado, usos, ultima_vez)
                VALUES (?,?,?,?,?,?,1,?)""",
                (firma, tipo_doc,
                 json.dumps(columnas or [], ensure_ascii=False),
                 fmt_fecha or '',
                 json.dumps(prefijos_doc or [], ensure_ascii=False),
                 banco_detectado or '', ahora))
        conn.commit(); conn.close()
    except Exception:
        pass

def buscar_formato_pdf(nombre_archivo, n_columnas):
    """Devuelve dict con info del formato guardado, o None si no existe."""
    if not OFFLINE_MODE:
        return None
    try:
        firma = _firma_pdf(nombre_archivo, n_columnas)
        conn  = _init_db()
        row   = conn.execute(
            """SELECT tipo_doc, columnas, fmt_fecha, prefijos_doc,
                      banco_detectado, usos, ultima_vez
               FROM pdf_formatos WHERE firma=?""", (firma,)).fetchone()
        conn.close()
        if not row:
            return None
        return {
            'tipo_doc'        : row[0],
            'columnas'        : json.loads(row[1] or '[]'),
            'fmt_fecha'       : row[2],
            'prefijos_doc'    : json.loads(row[3] or '[]'),
            'banco_detectado' : row[4],
            'usos'            : row[5],
            'ultima_vez'      : row[6],
        }
    except Exception:
        return None

def listar_formatos_aprendidos():
    """Devuelve todos los formatos guardados en el catálogo."""
    try:
        if not os.path.exists(DB_PATH): return []
        conn = _init_db()
        rows = conn.execute(
            """SELECT firma, tipo_doc, banco_detectado, usos, ultima_vez
               FROM pdf_formatos ORDER BY usos DESC""").fetchall()
        conn.close()
        return rows
    except Exception:
        return []

# ══════════════════════════════════════════════════════════════════════════════
# FASE D — Sistema de aprendizaje de conceptos NC
# ══════════════════════════════════════════════════════════════════════════════

# Stopwords para extraccion de tokens significativos
_STOP_NC = {
    'de','la','el','en','a','y','con','por','para','del','un','una','los','las',
    'al','se','su','que','no','es','pago','transferencia','nota','contable',
    'banco','bancario','cobro','cargo','por','desde','hasta','entre','sin',
    'mas','los','las','este','esta','fue','han','hay','bien','ser','tiene',
    'son','sus','les','nos','fue','era','ese','esa'
}

import hashlib as _hashlib
import unicodedata as _ud2

def _norm_nc(s):
    return _ud2.normalize('NFKD', (s or '').lower()).encode('ascii','ignore').decode()

def _extraer_tokens_nc(texto):
    """Extrae tokens significativos (3+ chars, sin stopwords) de un concepto NC."""
    norm = _norm_nc(texto)
    tokens = re.findall(r'[a-z0-9]{3,}', norm)
    return sorted(set(t for t in tokens if t not in _STOP_NC))

def _uuid_par_nc(banco_tokens, aux_tokens):
    """UUID determinista del par banco<->auxiliar (MD5 de tokens ordenados)."""
    key = '|'.join(sorted(banco_tokens)) + '::' + '|'.join(sorted(aux_tokens))
    return _hashlib.md5(key.encode()).hexdigest()[:16]

def _similitud_tokens_nc(t1, t2):
    """Similitud Jaccard entre dos listas de tokens."""
    s1, s2 = set(t1), set(t2)
    if not s1 or not s2:
        return 0.0
    return len(s1 & s2) / len(s1 | s2)

def buscar_en_catalogo_nc(banco_desc, aux_concepto, umbral=0.30):
    """
    Busca si el par banco<->NC tiene una regla aprobada.
    Devuelve (uuid, similitud) o (None, 0.0).
    """
    try:
        banco_tok = _extraer_tokens_nc(banco_desc)
        aux_tok   = _extraer_tokens_nc(aux_concepto)
        if len(banco_tok) < 1 or len(aux_tok) < 1:
            return None, 0.0
        conn  = _init_db()
        rows  = conn.execute(
            "SELECT uuid, banco_tokens, aux_tokens, confirmaciones FROM nc_catalogo "
            "WHERE nivel IN ('ALTA','MEDIA') ORDER BY confirmaciones DESC LIMIT 200"
        ).fetchall()
        conn.close()
        mejor_sim, mejor_uuid = 0.0, None
        for uuid, bt_j, at_j, _ in rows:
            bt = json.loads(bt_j or '[]')
            at = json.loads(at_j or '[]')
            sim = (_similitud_tokens_nc(banco_tok, bt) +
                   _similitud_tokens_nc(aux_tok,   at)) / 2
            if sim > mejor_sim and sim >= umbral:
                mejor_sim, mejor_uuid = sim, uuid
        return mejor_uuid, mejor_sim
    except Exception:
        return None, 0.0

def _promover_candidatos_nc(min_veces=3):
    """Promueve candidatos con suficientes confirmaciones al catalogo."""
    try:
        conn = _init_db()
        candidatos = conn.execute(
            "SELECT uuid, banco_desc_raw, aux_concepto_raw, banco_tokens, "
            "aux_tokens, veces_visto, fecha_primera, fecha_ultima "
            "FROM nc_aprendizaje WHERE veces_visto >= ?", (min_veces,)
        ).fetchall()
        n = 0
        for uuid, bd, ac, bt, at, vv, fp, fu in candidatos:
            nivel = 'ALTA' if vv >= 5 else 'MEDIA'
            existe = conn.execute(
                "SELECT id FROM nc_catalogo WHERE uuid=?", (uuid,)).fetchone()
            if not existe:
                conn.execute("""INSERT INTO nc_catalogo
                    (uuid, banco_tokens, aux_tokens, confirmaciones, nivel,
                     aprobado_por, fecha_primera, fecha_ultima, sync_status)
                    VALUES (?,?,?,?,?,'AUTO',?,?,'PENDIENTE_SYNC')""",
                    (uuid, bt, at, vv, nivel, fp, fu))
                n += 1
            conn.execute("DELETE FROM nc_aprendizaje WHERE uuid=?", (uuid,))
        conn.commit(); conn.close()
        return n
    except Exception:
        return 0

def _aprender_match_nc(banco_desc, aux_doc, aux_concepto, metodo,
                       valor_banco=None, valor_aux=None):
    """
    Registra un match NC confirmado y alimenta el aprendizaje.
    Solo aprende si el documento es NC- (notas contables).
    """
    if not aux_doc or not str(aux_doc).upper().startswith('NC-'):
        return
    try:
        banco_tok = _extraer_tokens_nc(banco_desc)
        aux_tok   = _extraer_tokens_nc(aux_concepto)
        if len(banco_tok) < 2 or len(aux_tok) < 2:
            return
        uuid  = _uuid_par_nc(banco_tok, aux_tok)
        ahora = datetime.now().isoformat(timespec='seconds')
        conn  = _init_db()
        # Log historico
        conn.execute("""INSERT INTO nc_historial_match
            (fecha, banco_desc, aux_doc, aux_concepto, metodo, valor_banco, valor_aux)
            VALUES (?,?,?,?,?,?,?)""",
            (ahora, (banco_desc or '')[:200], aux_doc,
             (aux_concepto or '')[:200], metodo, valor_banco, valor_aux))
        # Actualizar catalogo si ya existe la regla
        en_cat = conn.execute(
            "SELECT id, confirmaciones FROM nc_catalogo WHERE uuid=?", (uuid,)).fetchone()
        if en_cat:
            nuevo_nivel = 'ALTA' if en_cat[1]+1 >= 5 else 'MEDIA'
            conn.execute(
                "UPDATE nc_catalogo SET confirmaciones=?, nivel=?, "
                "fecha_ultima=?, sync_status='PENDIENTE_SYNC' WHERE uuid=?",
                (en_cat[1]+1, nuevo_nivel, ahora, uuid))
        else:
            # Actualizar o insertar en aprendizaje
            cand = conn.execute(
                "SELECT id, veces_visto FROM nc_aprendizaje WHERE uuid=?", (uuid,)).fetchone()
            if cand:
                conn.execute(
                    "UPDATE nc_aprendizaje SET veces_visto=?, fecha_ultima=? WHERE uuid=?",
                    (cand[1]+1, ahora, uuid))
            else:
                conn.execute("""INSERT INTO nc_aprendizaje
                    (uuid, banco_desc_raw, aux_concepto_raw, banco_tokens, aux_tokens,
                     veces_visto, fecha_primera, fecha_ultima)
                    VALUES (?,?,?,?,?,1,?,?)""",
                    (uuid, (banco_desc or '')[:200], (aux_concepto or '')[:200],
                     json.dumps(banco_tok), json.dumps(aux_tok), ahora, ahora))
        conn.commit(); conn.close()
        _promover_candidatos_nc()
    except Exception:
        pass

# ── Sync bidireccional SQLite <-> Google Sheets ───────────────────────────────

def _push_catalogo_to_sheets():
    """Sube reglas PENDIENTE_SYNC del SQLite a Google Sheets."""
    try:
        import gspread
        from google.oauth2.service_account import Credentials
        creds_json = st.secrets.get("GOOGLE_SHEETS_CREDS", None)
        sheet_id   = st.secrets.get("GOOGLE_SHEET_ID",    None)
        if not creds_json or not sheet_id:
            return 0
        creds = Credentials.from_service_account_info(
            json.loads(creds_json),
            scopes=["https://spreadsheets.google.com/feeds",
                    "https://www.googleapis.com/auth/drive"])
        gc = gspread.authorize(creds)
        wb = gc.open_by_key(sheet_id)
        try:
            ws = wb.worksheet("nc_catalogo")
        except Exception:
            ws = wb.add_worksheet("nc_catalogo", rows=2000, cols=9)
            ws.append_row(["uuid","banco_tokens","aux_tokens","confirmaciones",
                           "nivel","aprobado_por","fecha_primera","fecha_ultima"])
        conn = _init_db()
        pendientes = conn.execute(
            "SELECT uuid, banco_tokens, aux_tokens, confirmaciones, nivel, "
            "aprobado_por, fecha_primera, fecha_ultima "
            "FROM nc_catalogo WHERE sync_status='PENDIENTE_SYNC'"
        ).fetchall()
        existentes = {r[0] for r in ws.get_all_values()[1:] if r}
        n = 0
        for row in pendientes:
            if row[0] not in existentes:
                ws.append_row(list(row))
                n += 1
            conn.execute(
                "UPDATE nc_catalogo SET sync_status='SINCRONIZADO' WHERE uuid=?",
                (row[0],))
        conn.commit(); conn.close()
        return n
    except Exception:
        return 0

def _pull_catalogo_from_sheets():
    """Descarga reglas nuevas de Google Sheets al SQLite local."""
    try:
        import gspread
        from google.oauth2.service_account import Credentials
        creds_json = st.secrets.get("GOOGLE_SHEETS_CREDS", None)
        sheet_id   = st.secrets.get("GOOGLE_SHEET_ID",    None)
        if not creds_json or not sheet_id:
            return 0
        creds = Credentials.from_service_account_info(
            json.loads(creds_json),
            scopes=["https://spreadsheets.google.com/feeds",
                    "https://www.googleapis.com/auth/drive"])
        gc = gspread.authorize(creds)
        wb = gc.open_by_key(sheet_id)
        try:
            ws = wb.worksheet("nc_catalogo")
        except Exception:
            return 0
        rows = ws.get_all_values()
        if len(rows) <= 1:
            return 0
        conn = _init_db()
        n = 0
        for row in rows[1:]:
            if len(row) < 7 or not row[0]:
                continue
            uuid  = row[0]
            conf  = int(row[3]) if str(row[3]).isdigit() else 1
            nivel = row[4] or 'MEDIA'
            existe = conn.execute(
                "SELECT id, confirmaciones FROM nc_catalogo WHERE uuid=?", (uuid,)).fetchone()
            if not existe:
                conn.execute("""INSERT INTO nc_catalogo
                    (uuid, banco_tokens, aux_tokens, confirmaciones, nivel,
                     aprobado_por, fecha_primera, fecha_ultima, sync_status)
                    VALUES (?,?,?,?,?,?,?,?,'SINCRONIZADO')""",
                    (uuid, row[1], row[2], conf, nivel,
                     row[5] if len(row)>5 else 'AUTO',
                     row[6] if len(row)>6 else '',
                     row[7] if len(row)>7 else ''))
                n += 1
            elif conf > existe[1]:
                nuevo_nivel = 'ALTA' if conf >= 5 else 'MEDIA'
                conn.execute(
                    "UPDATE nc_catalogo SET confirmaciones=?, nivel=?, "
                    "sync_status='SINCRONIZADO' WHERE uuid=?",
                    (conf, nuevo_nivel, uuid))
        conn.commit(); conn.close()
        return n
    except Exception:
        return 0

def sincronizar_catalogo_nc():
    """Sincronizacion bidireccional: push local -> Sheets, pull Sheets -> local."""
    if not OFFLINE_MODE:
        return 0, 0
    n_up   = _push_catalogo_to_sheets()
    n_down = _pull_catalogo_from_sheets()
    return n_up, n_down

def listar_catalogo_nc(limite=8):
    """Devuelve las reglas del catalogo ordenadas por confirmaciones."""
    try:
        if not os.path.exists(DB_PATH):
            return [], 0
        conn  = _init_db()
        total = conn.execute("SELECT COUNT(*) FROM nc_catalogo").fetchone()[0]
        rows  = conn.execute(
            "SELECT uuid, banco_tokens, aux_tokens, confirmaciones, nivel, "
            "aprobado_por, fecha_ultima "
            "FROM nc_catalogo ORDER BY confirmaciones DESC LIMIT ?", (limite,)
        ).fetchall()
        pend  = conn.execute(
            "SELECT COUNT(*) FROM nc_aprendizaje").fetchone()[0]
        conn.close()
        return rows, total, pend
    except Exception:
        return [], 0, 0

def _aprender_match_nc_cloud(banco_desc, aux_doc, aux_concepto, metodo):
    """
    Para Streamlit Cloud: registra el aprendizaje NC directamente en Google Sheets
    (tabla nc_aprendizaje del spreadsheet).
    """
    if not aux_doc or not str(aux_doc).upper().startswith('NC-'):
        return
    try:
        banco_tok = _extraer_tokens_nc(banco_desc)
        aux_tok   = _extraer_tokens_nc(aux_concepto)
        if len(banco_tok) < 2 or len(aux_tok) < 2:
            return
        uuid  = _uuid_par_nc(banco_tok, aux_tok)
        ahora = datetime.now().isoformat(timespec='seconds')
        import gspread
        from google.oauth2.service_account import Credentials
        creds_json = st.secrets.get("GOOGLE_SHEETS_CREDS", None)
        sheet_id   = st.secrets.get("GOOGLE_SHEET_ID",    None)
        if not creds_json or not sheet_id:
            return
        creds = Credentials.from_service_account_info(
            json.loads(creds_json),
            scopes=["https://spreadsheets.google.com/feeds",
                    "https://www.googleapis.com/auth/drive"])
        gc = gspread.authorize(creds)
        wb = gc.open_by_key(sheet_id)
        try:
            ws = wb.worksheet("nc_aprendizaje")
        except Exception:
            ws = wb.add_worksheet("nc_aprendizaje", rows=2000, cols=8)
            ws.append_row(["uuid","banco_desc_raw","aux_concepto_raw",
                           "banco_tokens","aux_tokens","veces_visto",
                           "fecha_primera","fecha_ultima"])
        rows = ws.get_all_values()
        uuid_map = {r[0]: i+2 for i,r in enumerate(rows[1:]) if r and r[0]}
        if uuid in uuid_map:
            row_n = uuid_map[uuid]
            try:
                vv = int(ws.cell(row_n, 6).value or 1)
            except Exception:
                vv = 1
            ws.update_cell(row_n, 6, vv + 1)
            ws.update_cell(row_n, 8, ahora)
            # Si llega a 3 -> promover en nc_catalogo sheet
            if vv + 1 >= 3:
                _promover_cloud_to_catalogo(ws, row_n, uuid, banco_tok,
                                            aux_tok, vv+1, wb)
        else:
            ws.append_row([uuid, (banco_desc or '')[:150],
                           (aux_concepto or '')[:150],
                           json.dumps(banco_tok), json.dumps(aux_tok),
                           1, ahora, ahora])
    except Exception:
        pass

def _promover_cloud_to_catalogo(ws_aprendizaje, row_n, uuid, banco_tok,
                                  aux_tok, vv, wb):
    """Promueve un candidato al nc_catalogo en Google Sheets."""
    try:
        try:
            ws_cat = wb.worksheet("nc_catalogo")
        except Exception:
            ws_cat = wb.add_worksheet("nc_catalogo", rows=2000, cols=9)
            ws_cat.append_row(["uuid","banco_tokens","aux_tokens","confirmaciones",
                               "nivel","aprobado_por","fecha_primera","fecha_ultima"])
        existentes = {r[0] for r in ws_cat.get_all_values()[1:] if r}
        if uuid not in existentes:
            ahora = datetime.now().isoformat(timespec='seconds')
            nivel = 'ALTA' if vv >= 5 else 'MEDIA'
            ws_cat.append_row([uuid, json.dumps(banco_tok), json.dumps(aux_tok),
                               vv, nivel, 'AUTO', ahora, ahora])
    except Exception:
        pass

def registrar_aprendizaje_nc(banco_desc, aux_doc, aux_concepto, metodo,
                              valor_banco=None, valor_aux=None):
    """Punto de entrada unificado: SQLite (offline) o Sheets (cloud)."""
    if OFFLINE_MODE:
        _aprender_match_nc(banco_desc, aux_doc, aux_concepto, metodo,
                           valor_banco, valor_aux)
    else:
        _aprender_match_nc_cloud(banco_desc, aux_doc, aux_concepto, metodo)


# ── Helpers originales ────────────────────────────────────────────────────────
def cop(v):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return '                 N/A'
    signo = '-' if v < 0 else ' '
    return f'{signo}$ {abs(v):>18,.2f}'

def pct_bar(p, width=20):
    filled = int(p / 100 * width)
    return '[' + '█' * filled + '░' * (width - filled) + ']'

# ── Función OCR ─────────────────────────────────────────────────────────────
def ocr_pdf_page(pdf_path, page_number):
    """Devuelve el texto de una página específica usando OCR."""
    if not OCR_AVAILABLE:
        return ""
    try:
        images = convert_from_path(pdf_path, first_page=page_number, last_page=page_number)
        if images:
            return pytesseract.image_to_string(images[0], lang='spa')
    except Exception as e:
        logging.warning(f"Error OCR en página {page_number}: {e}")
    return ""

# ── Diagnóstico de legibilidad (original + OCR) ─────────────────────────────
def diagnosticar_pdf(ruta, tipo):
    resultado = {
        'archivo': ruta, 'tipo': tipo,
        'paginas_total': 0, 'paginas_con_texto': 0, 'paginas_sin_texto': 0,
        'total_chars': 0, 'total_words': 0, 'lineas_doc_encontradas': 0,
        'pct_paginas_legibles': 0.0, 'pct_estimado_datos': 0.0,
        'calidad': '', 'advertencias': [], 'ocr_usado': False
    }

    pat_doc = re.compile(r'(?:CON|CE|CG|NC|RE|RG)-\d+')
    pat_mov_banco = re.compile(r'^\d{1,2}/\d{2}\s+', re.MULTILINE)

    try:
        with pdfplumber.open(ruta) as pdf:
            resultado['paginas_total'] = len(pdf.pages)
            for pag in pdf.pages:
                texto = pag.extract_text() or ''
                if len(texto.strip()) > 30:
                    resultado['paginas_con_texto'] += 1
                    resultado['total_chars'] += len(texto)
                    resultado['total_words'] += len(texto.split())
                    if tipo == 'AUXILIAR':
                        resultado['lineas_doc_encontradas'] += len(pat_doc.findall(texto))
                    else:
                        resultado['lineas_doc_encontradas'] += len(pat_mov_banco.findall(texto))
                else:
                    # Intentar OCR
                    if OCR_AVAILABLE:
                        ocr_text = ocr_pdf_page(ruta, pag.page_number)
                        if len(ocr_text.strip()) > 30:
                            resultado['paginas_con_texto'] += 1
                            resultado['total_chars'] += len(ocr_text)
                            resultado['total_words'] += len(ocr_text.split())
                            resultado['advertencias'].append(
                                f'Pág. {pag.page_number}: texto extraído con OCR')
                            resultado['ocr_usado'] = True
                            if tipo == 'AUXILIAR':
                                resultado['lineas_doc_encontradas'] += len(pat_doc.findall(ocr_text))
                            else:
                                resultado['lineas_doc_encontradas'] += len(pat_mov_banco.findall(ocr_text))
                        else:
                            resultado['paginas_sin_texto'] += 1
                            resultado['advertencias'].append(
                                f'Pág. {pag.page_number}: sin texto (imagen sin OCR disponible o ilegible)')
                    else:
                        resultado['paginas_sin_texto'] += 1
                        resultado['advertencias'].append(
                            f'Pág. {pag.page_number}: sin texto (imagen escaneada y OCR no instalado)')
    except Exception as e:
        resultado['advertencias'].append(f'Error al abrir: {e}')
        return resultado

    n_tot = resultado['paginas_total']
    n_ok  = resultado['paginas_con_texto']
    resultado['pct_paginas_legibles'] = (n_ok / n_tot * 100) if n_tot > 0 else 0

    if resultado['lineas_doc_encontradas'] > 0:
        pct_datos = min(100, resultado['pct_paginas_legibles'] * 0.98 +
                        min(2, resultado['lineas_doc_encontradas'] / 10))
    else:
        pct_datos = resultado['pct_paginas_legibles'] * 0.5

    resultado['pct_estimado_datos'] = round(pct_datos, 1)
    if pct_datos >= 95:
        resultado['calidad'] = '🟢 EXCELENTE'
    elif pct_datos >= 80:
        resultado['calidad'] = '🟡 BUENA'
    elif pct_datos >= 50:
        resultado['calidad'] = '🟠 PARCIAL'
    else:
        resultado['calidad'] = '🔴 BAJA'
    return resultado

# ── Parseo banco original (PDF) ─────────────────────────────────────────────
def limpiar_num(t):
    t = str(t or '').strip()
    if not t:
        return None
    neg = t.startswith('-') or (t.startswith('(') and t.endswith(')'))
    t = re.sub(r'[\$\(\)\s]', '', t).replace(',', '')
    try:
        v = float(t)
        return -abs(v) if neg else v
    except:
        return None

def es_fecha_banco(t):
    return bool(re.match(r'^\d{1,2}/\d{2}$', str(t or '').strip()))

def parsear_banco_pdf(ruta, usar_ocr=False):
    registros = []
    resumen = {}
    pat_res = {
        'SALDO_ANTERIOR': r'SALDO\s+ANTERIOR\s+\$?\s*([\d,\.]+)',
        'TOTAL_ABONOS'  : r'TOTAL\s+ABONOS\s+\$?\s*([\d,\.]+)',
        'TOTAL_CARGOS'  : r'TOTAL\s+CARGOS\s+\$?\s*([\d,\.]+)',
        'SALDO_ACTUAL'  : r'SALDO\s+ACTUAL\s+\$?\s*([\d,\.]+)',
    }

    anio_extracto = datetime.now().year   # se refina en la 1ª página del PDF

    with pdfplumber.open(ruta) as pdf:
        for n_pag, pag in enumerate(pdf.pages):
            # Obtener texto, con fallback a OCR si está vacío y usar_ocr=True
            texto = pag.extract_text() or ''
            if len(texto.strip()) <= 30 and usar_ocr and OCR_AVAILABLE:
                texto = ocr_pdf_page(ruta, pag.page_number)

            if n_pag == 0:
                # Detectar año del período en la primera página (evita hardcoding)
                _m_anio = re.search(r'\b(20\d{2})\b', texto)
                if _m_anio:
                    anio_extracto = int(_m_anio.group(1))
                for clave, pat in pat_res.items():
                    m = re.search(pat, texto, re.IGNORECASE)
                    if m and clave not in resumen:
                        v = m.group(1).replace(',', '')
                        resumen[clave] = limpiar_num(v)

            tabla = pag.extract_table({
                'vertical_strategy': 'lines',
                'horizontal_strategy': 'lines',
            })
            if not tabla:
                tabla = pag.extract_table()

            if tabla:
                for fila in tabla:
                    if not fila:
                        continue
                    celdas = [str(c or '').strip() for c in fila]
                    fecha_raw = next((c for c in celdas if es_fecha_banco(c)), None)
                    if not fecha_raw:
                        continue
                    nums = []
                    for c in reversed(celdas):
                        v = limpiar_num(c)
                        if v is not None:
                            nums.insert(0, v)
                        elif nums:
                            break
                    if len(nums) < 1:
                        continue
                    saldo = nums[-1]
                    valor = nums[-2] if len(nums) >= 2 else None
                    idx_f = celdas.index(fecha_raw)
                    n_num = len(nums)
                    desc = ' '.join(c for c in celdas[idx_f+1:len(celdas)-n_num]
                                     if c and not es_fecha_banco(c))
                    desc = re.sub(r'\s+', ' ', desc).strip()
                    registros.append({
                        'FECHA_RAW': fecha_raw,
                        'FECHA': pd.to_datetime(f'{anio_extracto}/' + fecha_raw,
                                        format='%Y/%d/%m', errors='coerce'),
                        'DESCRIPCION': desc,
                        'VALOR': valor,
                        'SALDO': saldo,
                        'TIPO': 'ABONO' if (valor or 0) >= 0 else 'CARGO',
                        'PAGINA': n_pag + 1,
                    })
            else:
                # Fallback línea a línea usando el texto (ya sea normal u OCR)
                for linea in texto.split('\n'):
                    partes = linea.strip().split()
                    if not partes or not es_fecha_banco(partes[0]):
                        continue
                    fecha_raw = partes[0]
                    nums = []; desc_p = []
                    for p in partes[1:]:
                        v = limpiar_num(p)
                        if v is not None:
                            nums.append(v)
                        elif not nums:
                            desc_p.append(p)
                    if not nums:
                        continue
                    saldo = nums[-1]
                    valor = nums[-2] if len(nums) >= 2 else nums[0]
                    registros.append({
                        'FECHA_RAW': fecha_raw,
                        'FECHA': pd.to_datetime(f'{anio_extracto}/' + fecha_raw,
                                        format='%Y/%d/%m', errors='coerce'),
                        'DESCRIPCION': ' '.join(desc_p),
                        'VALOR': valor,
                        'SALDO': saldo,
                        'TIPO': 'ABONO' if (valor or 0) >= 0 else 'CARGO',
                        'PAGINA': n_pag + 1,
                    })

    df = pd.DataFrame(registros)
    if df.empty:
        return df, resumen
    df = df[df['VALOR'].notna()].copy()
    df['VALOR'] = pd.to_numeric(df['VALOR'], errors='coerce')
    df = df.drop_duplicates(subset=['FECHA_RAW', 'DESCRIPCION', 'VALOR', 'SALDO'])
    df = df.sort_values('FECHA', na_position='last').reset_index(drop=True)
    df.index += 1
    return df, resumen

# ── Parseo auxiliar original (PDF) + OCR ─────────────────────────────────────
# ── Reglas por concepto (se evalúan ANTES del prefijo de documento) ──────────
# Orden importa: más específico primero
REGLAS_COL = [
    # ── DÉBITOS (entradas a la cuenta) ────────────────────────────────────────
    (re.compile(r'ABONO\s+A\s+PRESTAMO',             re.I), 'DEBITO'),
    (re.compile(r'RENDIMIENTO|INTERES\s+AHORROS',     re.I), 'DEBITO'),
    (re.compile(r'RECAUDO|INGRESO\s+CAJA|CONSIGNACI', re.I), 'DEBITO'),
    (re.compile(r'ABONO\s+CARTERA|ABONO\s+CUENTA',   re.I), 'DEBITO'),
    (re.compile(r'\bN\.D\.\b',                     re.I), 'DEBITO'),
    # ── CRÉDITOS (salidas / cargos bancarios) ─────────────────────────────────
    (re.compile(r'COMISION|COBRO\s+IVA|IVA\s+PAGOS', re.I), 'CREDITO'),
    (re.compile(r'4\s*POR\s*MIL|IMPTO\s+GOB|GRAVAMEN', re.I), 'CREDITO'),
    (re.compile(r'NOTA\s+CONTABLE|CARGO\s+BANC',     re.I), 'CREDITO'),
    (re.compile(r'NEQUI|PSE|DAVIPLATA|TRANSFIYA',      re.I), 'CREDITO'),
    (re.compile(r'\bPRESTAMO\b(?!.*ABONO)',           re.I), 'CREDITO'),
    (re.compile(r'RETIRO\s+PARA\s+PAGO',             re.I), 'CREDITO'),
    (re.compile(r'CANCELACION\s+NOMINA',              re.I), 'CREDITO'),
    (re.compile(r'GASTO\s+BANCAR|\bN\.C\.\b',     re.I), 'CREDITO'),
    (re.compile(r'IMPUESTO\s+MOVIMIENTO|GMF|4X1000',  re.I), 'CREDITO'),
    (re.compile(r'ND\s+POR\s+RECHAZO|RECHAZO\s+PAGO', re.I), 'CREDITO'),
    (re.compile(r'CUOTA\s+CREDITO|CUOTA\s+PRESTAMO', re.I), 'CREDITO'),
]

def determinar_columna(concepto, doc_code):
    """
    Determina si un asiento va a DÉBITO o CRÉDITO.
    Prioridad: 1) Concepto (REGLAS_COL)  2) Prefijo del documento
    Prefijos conocidos:
        CE- Comprobante Egreso     → CRÉDITO  (pago a proveedor)
        NC- Nota Contable          → CRÉDITO  (comisión/cargo bancario)
        CG- Comprobante Ingreso    → DÉBITO   (entrada de dinero)
        CON- Comprobante General   → DÉBITO   (por defecto)
        CO- igual que CON-         → DÉBITO
    """
    for pat, col in REGLAS_COL:
        if pat.search(concepto or ''):
            return col
    doc_prefix = (doc_code[:3].upper() if doc_code and len(doc_code) >= 3
                  else (doc_code or '')[:2].upper())
    # Prefijos de egreso / cargo
    if doc_prefix in ('CE-', 'NC-'):
        return 'CREDITO'
    if doc_prefix[:2] in ('CE', 'NC'):
        return 'CREDITO'
    # Prefijos de ingreso / abono
    if doc_prefix in ('CG-', 'CON'):
        return 'DEBITO'
    if doc_prefix[:2] in ('CG', 'CO'):
        return 'DEBITO'
    # Fallback conservador: la mayoría de asientos auxiliares son egresos
    return 'CREDITO'

def parsear_auxiliar_pdf(ruta, usar_ocr=False):
    texto_completo = ''
    n_pags_ok = 0
    n_pags_mal = 0
    meta = {}

    with pdfplumber.open(ruta) as pdf:
        for pag in pdf.pages:
            t = pag.extract_text() or ''
            if len(t.strip()) <= 30 and usar_ocr and OCR_AVAILABLE:
                t = ocr_pdf_page(ruta, pag.page_number)
            if len(t.strip()) > 30:
                n_pags_ok += 1
                texto_completo += '\n' + t
            else:
                n_pags_mal += 1

    m_si = re.search(r'Saldo\s+Inicial[:\s]+([\d,\.]+)', texto_completo, re.I)
    m_sf = re.search(r'Saldo\s+Final[:\s]+([\d,\.]+)', texto_completo, re.I)
    m_td = re.search(r'Subtotales.*?([\d]{1,3}(?:[,\\.][\d]{3})+(?:\.[\d]+)?)'
                     r'\s+([\d]{1,3}(?:[,\\.][\d]{3})+(?:\.[\d]+)?)',
                     texto_completo, re.I | re.DOTALL)
    meta['SALDO_INICIAL']  = limpiar_num((m_si.group(1) if m_si else '0').replace(',', ''))
    meta['SALDO_FINAL']    = limpiar_num((m_sf.group(1) if m_sf else '0').replace(',', ''))
    if m_td:
        meta['TOTAL_DEBITOS']  = limpiar_num(m_td.group(1).replace(',', ''))
        meta['TOTAL_CREDITOS'] = limpiar_num(m_td.group(2).replace(',', ''))
    else:
        meta['TOTAL_DEBITOS'] = meta['TOTAL_CREDITOS'] = 0
    meta['N_PAGS_OK']  = n_pags_ok
    meta['N_PAGS_MAL'] = n_pags_mal

    # ── Parseo línea a línea (mismo código original) ─────────────────────────
    PAT_DOC    = re.compile(r'^((?:CON|CE|CG|NC|RE|RG)-\d+)\s+(\d{1,2}/\d{1,2}/\d{4})\s+(.*)')
    PAT_MONTO  = re.compile(r'^([\d]{1,3}(?:,[\d]{3})*(?:\.[\d]{1,2})?)$')
    PAT_MPFX   = re.compile(r'^([\d]{1,3}(?:,[\d]{3})*(?:\.[\d]{1,2})?)\s+((?:CON|CE|CG|NC|RE|RG)-\d+.*)$')
    PAT_MSFX   = re.compile(r'\s([\d]{1,3}(?:,[\d]{3})*(?:\.[\d]{1,2})?)$')

    registros   = []
    lineas      = [l.strip() for l in texto_completo.split('\n') if l.strip()]
    pending_doc = None

    def guardar(doc, fecha_str, concepto, monto_str):
        monto = limpiar_num(monto_str.replace(',', ''))
        if not monto or monto <= 0:
            return
        col = determinar_columna(concepto, doc)
        debito  = monto if col == 'DEBITO'  else None
        credito = monto if col == 'CREDITO' else None
        try:
            fdt = pd.to_datetime(fecha_str, format='%d/%m/%Y', errors='coerce')
        except:
            fdt = pd.NaT
        registros.append({
            'DOCUMENTO' : doc,
            'FECHA_RAW' : fecha_str,
            'FECHA'     : fdt,
            'CONCEPTO'  : concepto,
            'DEBITO'    : debito,
            'CREDITO'   : credito,
            'COLUMNA'   : col,
            'VALOR_NETO': (debito or 0) - (credito or 0),
        })

    for linea in lineas:
        m_pfx = PAT_MPFX.match(linea)
        if m_pfx:
            monto_ant = m_pfx.group(1)
            resto     = m_pfx.group(2)
            if pending_doc:
                guardar(pending_doc['doc'], pending_doc['date'],
                        pending_doc['concept'], monto_ant)
                pending_doc = None
            m_doc = PAT_DOC.match(resto)
            if m_doc:
                doc_c   = m_doc.group(1)
                fecha_s = m_doc.group(2)
                concepto_raw = m_doc.group(3)
                m_end = PAT_MSFX.search(concepto_raw)
                if m_end:
                    monto_end = m_end.group(1)
                    concepto_limpio = concepto_raw[:m_end.start()].strip()
                    guardar(doc_c, fecha_s, concepto_limpio, monto_end)
                else:
                    pending_doc = {'doc': doc_c, 'date': fecha_s, 'concept': concepto_raw}
            continue

        m_doc = PAT_DOC.match(linea)
        if m_doc:
            if pending_doc:
                pass
            doc_c   = m_doc.group(1)
            fecha_s = m_doc.group(2)
            concepto_raw = m_doc.group(3)
            m_end = PAT_MSFX.search(concepto_raw)
            if m_end:
                monto_end = m_end.group(1)
                concepto_limpio = concepto_raw[:m_end.start()].strip()
                guardar(doc_c, fecha_s, concepto_limpio, monto_end)
            else:
                pending_doc = {'doc': doc_c, 'date': fecha_s, 'concept': concepto_raw}
            continue

        m_num = PAT_MONTO.match(linea)
        if m_num and pending_doc:
            guardar(pending_doc['doc'], pending_doc['date'],
                    pending_doc['concept'], m_num.group(1))
            pending_doc = None
            continue

        if pending_doc:
            m_end = PAT_MSFX.search(linea)
            if m_end:
                monto_end = m_end.group(1)
                try:
                    mval = float(monto_end.replace(',',''))
                    if mval > 100:
                        guardar(pending_doc['doc'], pending_doc['date'],
                                pending_doc['concept'], monto_end)
                        pending_doc = None
                except:
                    pass

    df = pd.DataFrame(registros)
    if not df.empty:
        df = df.drop_duplicates(subset=['DOCUMENTO','FECHA_RAW','DEBITO','CREDITO'])
        df = df.sort_values('FECHA', na_position='last').reset_index(drop=True)
        df.index += 1
    # Si los totales no se detectaron, calcular desde el dataframe
    if not meta['TOTAL_DEBITOS']:
        meta['TOTAL_DEBITOS'] = df['DEBITO'].sum() if not df.empty else 0
    if not meta['TOTAL_CREDITOS']:
        meta['TOTAL_CREDITOS'] = df['CREDITO'].sum() if not df.empty else 0
    # ── Fase C: aprender formato si el auxiliar fue parseado correctamente ──
    if not df.empty:
        prefijos_vistos = sorted(df['DOCUMENTO'].apply(_prefijo_doc).unique().tolist())                           if 'DOCUMENTO' in df.columns else []
        registrar_formato_pdf(
            nombre_archivo = meta.get('_nombre_archivo', ''),
            tipo_doc       = 'auxiliar',
            columnas       = list(df.columns),
            fmt_fecha      = 'DD/MM/YYYY',
            prefijos_doc   = [p for p in prefijos_vistos if p],
        )
    return df, meta

# ── Parseo para CSV/Excel/TXT ─────────────────────────────────────────────
def parsear_banco_csv(df):
    registros = []
    resumen = {}
    # Detectar columnas por nombre
    col_fecha = next((c for c in df.columns if 'fecha' in c.lower()), None)
    col_desc  = next((c for c in df.columns if 'descrip' in c.lower() or 'concepto' in c.lower()), None)
    col_valor = next((c for c in df.columns if 'valor' in c.lower() or 'monto' in c.lower()), None)
    col_saldo = next((c for c in df.columns if 'saldo' in c.lower()), None)

    if not col_fecha: raise ValueError("No se encontró columna de fecha en el archivo del banco")
    for _, row in df.iterrows():
        fecha = str(row[col_fecha])
        try:
            fecha_dt = pd.to_datetime(fecha, dayfirst=True, errors='coerce')
        except:
            fecha_dt = pd.NaT
        desc = str(row[col_desc]) if col_desc else ''
        valor = limpiar_num(row[col_valor]) if col_valor else 0
        saldo = limpiar_num(row[col_saldo]) if col_saldo else None
        registros.append({
            'FECHA_RAW': fecha, 'FECHA': fecha_dt,
            'DESCRIPCION': desc, 'VALOR': valor, 'SALDO': saldo,
            'TIPO': 'ABONO' if (valor or 0) >= 0 else 'CARGO'
        })
    df_out = pd.DataFrame(registros)
    df_out = df_out[df_out['VALOR'].notna()]
    df_out['VALOR'] = pd.to_numeric(df_out['VALOR'], errors='coerce')
    df_out = df_out.drop_duplicates()
    df_out = df_out.sort_values('FECHA', na_position='last').reset_index(drop=True)
    df_out.index += 1
    # Calcular totales
    resumen['TOTAL_ABONOS'] = df_out[df_out['VALOR'] > 0]['VALOR'].sum()
    resumen['TOTAL_CARGOS'] = df_out[df_out['VALOR'] < 0]['VALOR'].sum()
    if col_saldo:
        resumen['SALDO_INICIAL'] = df_out.iloc[0]['SALDO'] if not df_out.empty else 0
        resumen['SALDO_FINAL']   = df_out.iloc[-1]['SALDO'] if not df_out.empty else 0
    else:
        resumen['SALDO_INICIAL'] = resumen['SALDO_FINAL'] = 0
    return df_out, resumen

def _col(df, *palabras):
    """Busca columna por palabras clave (normaliza acentos para comparación)."""
    import unicodedata
    def norm(s):
        return unicodedata.normalize('NFKD', s.lower()).encode('ascii', 'ignore').decode()
    palabras_norm = [norm(p) for p in palabras]
    return next((c for c in df.columns if any(p in norm(c) for p in palabras_norm)), None)

def parsear_auxiliar_csv(df):
    registros = []
    meta = {}
    col_doc = _col(df, 'documento', 'doc')
    col_fec = _col(df, 'fecha')
    col_con = _col(df, 'concepto', 'descrip')
    col_deb = _col(df, 'debito', 'débito', 'debe', 'debitos', 'débitos')
    col_cre = _col(df, 'credito', 'crédito', 'haber', 'creditos', 'créditos')

    if not col_fec: raise ValueError("No se encontró columna de fecha en el auxiliar")

    # Filtrar filas que no sean movimientos reales (sin doc o sin fecha válida)
    for _, row in df.iterrows():
        doc = str(row[col_doc]).strip() if col_doc else ''
        fecha = str(row[col_fec]).strip()
        if not fecha or fecha in ('nan', 'NaT', 'Fecha'): continue
        try: fecha_dt = pd.to_datetime(fecha, dayfirst=True, errors='coerce')
        except: fecha_dt = pd.NaT
        if pd.isna(fecha_dt): continue
        concepto = str(row[col_con]).strip() if col_con else ''
        debito  = limpiar_num(row[col_deb])  if col_deb else None
        credito = limpiar_num(row[col_cre]) if col_cre else None
        # Ignorar filas de subtotales/encabezados de cuenta (sin doc real)
        if not re.match(r'^[A-Z]{2,3}-\d+', doc) and not doc:
            continue
        col_asiento = determinar_columna(concepto, doc)
        registros.append({
            'DOCUMENTO': doc, 'FECHA_RAW': fecha, 'FECHA': fecha_dt,
            'CONCEPTO': concepto, 'DEBITO': debito, 'CREDITO': credito,
            'COLUMNA': col_asiento, 'VALOR_NETO': (debito or 0) - (credito or 0)
        })
    df_out = pd.DataFrame(registros)
    if not df_out.empty:
        # CSV es exportación autoritativa del sistema contable; no deduplicar
        # (misma cuenta puede tener múltiples líneas idénticas en un comprobante)
        df_out = df_out.sort_values('FECHA', na_position='last').reset_index(drop=True)
        df_out.index += 1
    meta['TOTAL_DEBITOS']  = df_out['DEBITO'].sum()  if not df_out.empty else 0
    meta['TOTAL_CREDITOS'] = df_out['CREDITO'].sum() if not df_out.empty else 0
    # Buscar saldo inicial en el CSV (fila con 'Saldo Inicial:')
    if 'SALDO_INICIAL' not in meta:
        meta['SALDO_INICIAL'] = 0
    if 'SALDO_FINAL' not in meta:
        meta['SALDO_FINAL'] = meta['SALDO_INICIAL'] + meta['TOTAL_DEBITOS'] - meta['TOTAL_CREDITOS']
    return df_out, meta

def parsear_banco_txt(texto):
    registros = []
    resumen = {}
    # Detectar año en el texto (evita hardcoding)
    _m_anio_t = re.search(r'\b(20\d{2})\b', texto or '')
    anio_extracto = int(_m_anio_t.group(1)) if _m_anio_t else datetime.now().year
    for linea in texto.split('\n'):
        partes = linea.strip().split()
        if not partes or not es_fecha_banco(partes[0]): continue
        fecha_raw = partes[0]
        nums = []; desc_p = []
        for p in partes[1:]:
            v = limpiar_num(p)
            if v is not None: nums.append(v)
            elif not nums: desc_p.append(p)
        if not nums: continue
        saldo = nums[-1]
        valor = nums[-2] if len(nums) >= 2 else nums[0]
        registros.append({
            'FECHA_RAW': fecha_raw,
            'FECHA': pd.to_datetime(f'{anio_extracto}/' + fecha_raw, format='%Y/%d/%m', errors='coerce'),
            'DESCRIPCION': ' '.join(desc_p), 'VALOR': valor, 'SALDO': saldo,
            'TIPO': 'ABONO' if (valor or 0) >= 0 else 'CARGO'
        })
    df = pd.DataFrame(registros)
    if not df.empty:
        df = df[df['VALOR'].notna()]
        df['VALOR'] = pd.to_numeric(df['VALOR'], errors='coerce')
        df = df.drop_duplicates()
        df = df.sort_values('FECHA', na_position='last').reset_index(drop=True)
        df.index += 1
    resumen['TOTAL_ABONOS'] = df[df['VALOR'] > 0]['VALOR'].sum()
    resumen['TOTAL_CARGOS'] = df[df['VALOR'] < 0]['VALOR'].sum()
    resumen['SALDO_INICIAL'] = df.iloc[0]['SALDO'] if not df.empty else 0
    resumen['SALDO_FINAL']   = df.iloc[-1]['SALDO'] if not df.empty else 0
    return df, resumen

def parsear_auxiliar_txt(texto_completo):
    meta = {}
    registros = []
    lineas = [l.strip() for l in texto_completo.split('\n') if l.strip()]
    pending_doc = None
    PAT_DOC = re.compile(r'^((?:CON|CE|CG|NC|RE|RG)-\d+)\s+(\d{1,2}/\d{1,2}/\d{4})\s+(.*)')
    PAT_MONTO = re.compile(r'^([\d]{1,3}(?:,[\d]{3})*(?:\.[\d]{1,2})?)$')
    PAT_MPFX  = re.compile(r'^([\d]{1,3}(?:,[\d]{3})*(?:\.[\d]{1,2})?)\s+((?:CON|CE|CG|NC|RE|RG)-\d+.*)$')
    PAT_MSFX  = re.compile(r'\s([\d]{1,3}(?:,[\d]{3})*(?:\.[\d]{1,2})?)$')
    def guardar(doc, fecha_s, concepto, monto_str):
        monto = limpiar_num(monto_str.replace(',',''))
        if not monto or monto <= 0: return
        col = determinar_columna(concepto, doc)
        debito = monto if col=='DEBITO' else None
        credito = monto if col=='CREDITO' else None
        try: fdt = pd.to_datetime(fecha_s, format='%d/%m/%Y', errors='coerce')
        except: fdt = pd.NaT
        registros.append({
            'DOCUMENTO': doc, 'FECHA_RAW': fecha_s, 'FECHA': fdt,
            'CONCEPTO': concepto, 'DEBITO': debito, 'CREDITO': credito,
            'COLUMNA': col, 'VALOR_NETO': (debito or 0) - (credito or 0)
        })
    for linea in lineas:
        m_pfx = PAT_MPFX.match(linea)
        if m_pfx:
            if pending_doc:
                guardar(pending_doc['doc'], pending_doc['date'], pending_doc['concept'], m_pfx.group(1))
                pending_doc = None
            m_doc = PAT_DOC.match(m_pfx.group(2))
            if m_doc:
                doc_c, fecha_s, concepto_raw = m_doc.group(1), m_doc.group(2), m_doc.group(3)
                m_end = PAT_MSFX.search(concepto_raw)
                if m_end:
                    guardar(doc_c, fecha_s, concepto_raw[:m_end.start()].strip(), m_end.group(1))
                else:
                    pending_doc = {'doc': doc_c, 'date': fecha_s, 'concept': concepto_raw}
            continue
        m_doc = PAT_DOC.match(linea)
        if m_doc:
            doc_c, fecha_s, concepto_raw = m_doc.group(1), m_doc.group(2), m_doc.group(3)
            m_end = PAT_MSFX.search(concepto_raw)
            if m_end:
                guardar(doc_c, fecha_s, concepto_raw[:m_end.start()].strip(), m_end.group(1))
            else:
                pending_doc = {'doc': doc_c, 'date': fecha_s, 'concept': concepto_raw}
            continue
        if PAT_MONTO.match(linea) and pending_doc:
            guardar(pending_doc['doc'], pending_doc['date'], pending_doc['concept'], linea)
            pending_doc = None
            continue
        if pending_doc:
            m_end = PAT_MSFX.search(linea)
            if m_end:
                try:
                    if float(m_end.group(1).replace(',','')) > 100:
                        guardar(pending_doc['doc'], pending_doc['date'], pending_doc['concept'], m_end.group(1))
                        pending_doc = None
                except: pass
    df = pd.DataFrame(registros)
    df = df.drop_duplicates()
    df = df.sort_values('FECHA', na_position='last').reset_index(drop=True)
    df.index += 1
    meta['TOTAL_DEBITOS'] = df['DEBITO'].sum() if not df.empty else 0
    meta['TOTAL_CREDITOS'] = df['CREDITO'].sum() if not df.empty else 0
    # Saldos no disponibles en TXT generalmente
    meta['SALDO_INICIAL'] = meta['SALDO_FINAL'] = 0
    return df, meta

# ══════════════════════════════════════════════════════════════════════════════
# REGISTRO EXTENSIBLE DE FORMATOS
# ──────────────────────────────────────────────────────────────────────────────
# Cómo añadir un nuevo formato:
#   1. Escribir fn_detectar(ruta, muestra_texto) -> float  [0.0–1.0 confianza]
#   2. Escribir fn_parsear(ruta, usar_ocr)       -> (DataFrame, dict_meta)
#   3. Agregar una entrada al final de REGISTRO_FORMATOS con tipo/ext correctos.
#   No tocar nada más.
# ══════════════════════════════════════════════════════════════════════════════

import unicodedata as _ud

def _norm(s):
    return _ud.normalize('NFKD', s.lower()).encode('ascii', 'ignore').decode()

def _muestra_texto(ruta, ext, n_lineas=50):
    """Texto de muestra para detección rápida (sin parsear el archivo completo)."""
    try:
        if ext == '.pdf':
            with pdfplumber.open(ruta) as pdf:
                return (pdf.pages[0].extract_text() or '') if pdf.pages else ''
        else:
            with open(ruta, 'r', encoding='latin1', errors='replace') as f:
                return ''.join(f.readlines()[:n_lineas])
    except Exception:
        return ''

def _header_row_csv(ruta, encoding='latin1'):
    """Fila donde empieza el encabezado real del CSV (salta metadatos)."""
    claves = {'documento', 'fecha', 'concepto', 'debito', 'credito',
              'valor', 'saldo', 'descripcion'}
    with open(ruta, 'r', encoding=encoding, errors='replace') as f:
        for i, linea in enumerate(f):
            hits = sum(1 for k in claves if k in _norm(linea))
            if hits >= 3:
                return i
    return 0

def _leer_csv_inteligente(ruta):
    """Lee CSV saltando metadatos y capturando saldo inicial si existe."""
    skip = _header_row_csv(ruta)
    saldo_ini = None
    with open(ruta, 'r', encoding='latin1', errors='replace') as f:
        for linea in f.readlines()[:skip + 3]:
            m = re.search(r'Saldo\s+Inicial[:\s]+([\d,\.]+)', linea, re.I)
            if m:
                saldo_ini = limpiar_num(m.group(1).replace(',', ''))
                break
    df = pd.read_csv(ruta, encoding='latin1', sep=None, engine='python',
                     skiprows=skip, header=0)
    df = df.dropna(how='all').reset_index(drop=True)
    return df, saldo_ini

# ── Detectores ────────────────────────────────────────────────────────────────

def _det_bancolombia_pdf(ruta, m):
    hits = [
        bool(re.search(r'ESTADO\s+DE\s+CUENTA',      m, re.I)),
        bool(re.search(r'SALDO\s+ANTERIOR',           m, re.I)),
        bool(re.search(r'TOTAL\s+ABONOS',             m, re.I)),
        bool(re.search(r'TOTAL\s+CARGOS',             m, re.I)),
        bool(re.search(r'BANCOLOMBIA',                m, re.I)),
    ]
    return sum(hits) / len(hits)

def _det_siigo_aux_csv(ruta, m):
    hits = [
        bool(re.search(r'Auxiliares\s*[-–]\s*Plan\s+de\s+Cuentas', m, re.I)),
        bool(re.search(r'Nit\.?\s*\d{6,}',            m, re.I)),
        bool(re.search(r'(?:CE|CON|NC)-\d+',          m)),
        bool(re.search(r'D[eé]bitos?',                m, re.I)),
        bool(re.search(r'Cr[eé]ditos?',               m, re.I)),
        bool(re.search(r'Saldo\s+Inicial',            m, re.I)),
    ]
    return sum(hits) / len(hits)

def _det_helisa_aux_csv(ruta, m):
    hits = [
        bool(re.search(r'HELISA',                     m, re.I)),
        bool(re.search(r'Libro\s+Auxiliar',           m, re.I)),
        bool(re.search(r'(?:CE|CON|NC|RE|RG)-\d+',   m)),
        bool(re.search(r'D[eé]bito|Cr[eé]dito',      m, re.I)),
    ]
    return sum(hits) / len(hits)

def _det_world_office_aux_csv(ruta, m):
    hits = [
        bool(re.search(r'World\s*Office|WO\s+\d',    m, re.I)),
        bool(re.search(r'Comprobante',                m, re.I)),
        bool(re.search(r'D[eé]bito|Cr[eé]dito',      m, re.I)),
        bool(re.search(r'\d{4}-\d{2}-\d{2}',         m)),
    ]
    return sum(hits) / len(hits)

def _det_siigo_aux_pdf(ruta, m):
    """Detector especifico para exportaciones PDF de SIIGO Auxiliares de Plan de Cuentas."""
    hits = [
        bool(re.search(r'(?:CON|CE|CG|NC|RE|RG)-\d+',    m)),         # codigos de documento
        bool(re.search(r'D[e\xe9]bitos?',                  m, re.I)),   # columna Debitos
        bool(re.search(r'Cr[e\xe9]ditos?',                 m, re.I)),   # columna Creditos
        bool(re.search(r'Saldo\s+Inicial|Saldo\s+Final',  m, re.I)),   # saldos
        bool(re.search(r'Auxiliares|Plan\s+de\s+Cuentas', m, re.I)),   # encabezado SIIGO
        bool(re.search(r'\d{1,2}/\d{1,2}/\d{4}',         m)),         # formato fecha DD/MM/YYYY
    ]
    return sum(hits) / len(hits)

def _det_aux_pdf_generico(ruta, m):
    """Detector generico para cualquier auxiliar en PDF con prefijos de documento."""
    hits = [
        bool(re.search(r'(?:CON|CE|CG|NC|RE|RG)-\d+',    m)),
        bool(re.search(r'D[e\xe9]bitos?',                  m, re.I)),
        bool(re.search(r'Cr[e\xe9]ditos?',                 m, re.I)),
        bool(re.search(r'Saldo',                           m, re.I)),
        bool(re.search(r'\d{1,2}/\d{1,2}/\d{4}',         m)),
    ]
    return sum(hits) / len(hits)

def _det_banco_csv_generico(ruta, m):
    hits = [
        bool(re.search(r'\bfecha\b',                  m, re.I)),
        bool(re.search(r'\bvalor\b|\bmonto\b',        m, re.I)),
        bool(re.search(r'\bsaldo\b',                  m, re.I)),
    ]
    return sum(hits) / len(hits)

def _det_aux_csv_generico(ruta, m):
    hits = [
        bool(re.search(r'\bdocumento\b',              m, re.I)),
        bool(re.search(r'\bfecha\b',                  m, re.I)),
        bool(re.search(r'\bconcepto\b|\bdescripci',   m, re.I)),
        bool(re.search(r'\bd[eé]bito\b|\bhaber\b',   m, re.I)),
    ]
    return sum(hits) / len(hits)

def _det_banco_txt(ruta, m):
    lineas_fecha = sum(1 for l in m.split('\n')
                       if re.match(r'^\d{1,2}/\d{2}\s', l.strip()))
    return min(1.0, lineas_fecha / 5)

def _det_aux_txt(ruta, m):
    hits = [
        bool(re.search(r'(?:CON|CE|NC)-\d+',         m)),
        bool(re.search(r'\d{1,2}/\d{2}/\d{4}',       m)),
    ]
    return sum(hits) / len(hits)

# ── Parsers de formato ─────────────────────────────────────────────────────────

def _par_bancolombia_pdf(ruta, usar_ocr):
    return parsear_banco_pdf(ruta, usar_ocr=usar_ocr)

def _par_siigo_aux_csv(ruta, usar_ocr):
    df_raw, saldo_ini = _leer_csv_inteligente(ruta)
    df, meta = parsear_auxiliar_csv(df_raw)
    if saldo_ini is not None:
        meta['SALDO_INICIAL'] = saldo_ini
        meta['SALDO_FINAL'] = saldo_ini + meta.get('TOTAL_DEBITOS',0) - meta.get('TOTAL_CREDITOS',0)
    return df, meta

def _par_helisa_aux_csv(ruta, usar_ocr):
    # Helisa exporta similar a SIIGO; reutilizamos misma lógica
    return _par_siigo_aux_csv(ruta, usar_ocr)

def _par_world_office_aux_csv(ruta, usar_ocr):
    # World Office: columnas pueden llamarse "Debe"/"Haber" en lugar de Débito/Crédito
    df_raw, saldo_ini = _leer_csv_inteligente(ruta)
    df, meta = parsear_auxiliar_csv(df_raw)
    if saldo_ini is not None:
        meta['SALDO_INICIAL'] = saldo_ini
        meta['SALDO_FINAL'] = saldo_ini + meta.get('TOTAL_DEBITOS',0) - meta.get('TOTAL_CREDITOS',0)
    return df, meta

def _par_aux_pdf_generico(ruta, usar_ocr):
    return parsear_auxiliar_pdf(ruta, usar_ocr=usar_ocr)

def _par_banco_csv_generico(ruta, usar_ocr):
    df_raw, _ = _leer_csv_inteligente(ruta)
    return parsear_banco_csv(df_raw)

def _par_aux_csv_generico(ruta, usar_ocr):
    df_raw, saldo_ini = _leer_csv_inteligente(ruta)
    df, meta = parsear_auxiliar_csv(df_raw)
    if saldo_ini is not None:
        meta['SALDO_INICIAL'] = saldo_ini
        meta['SALDO_FINAL'] = saldo_ini + meta.get('TOTAL_DEBITOS',0) - meta.get('TOTAL_CREDITOS',0)
    return df, meta

def _par_banco_txt(ruta, usar_ocr):
    with open(ruta, 'r', encoding='utf-8', errors='replace') as f:
        return parsear_banco_txt(f.read())

def _par_aux_txt(ruta, usar_ocr):
    with open(ruta, 'r', encoding='utf-8', errors='replace') as f:
        return parsear_auxiliar_txt(f.read())

# ── Registro principal ────────────────────────────────────────────────────────
# Orden importa: formatos más específicos primero dentro del mismo tipo/ext.
# El despachador elige el de mayor confianza, no el primero.

REGISTRO_FORMATOS = [
    # ── BANCO ─────────────────────────────────────────────────────────────────
    {
        'nombre'  : 'Bancolombia — PDF Estado de Cuenta',
        'tipo'    : 'BANCO',
        'ext'     : ['.pdf'],
        'detectar': _det_bancolombia_pdf,
        'parsear' : _par_bancolombia_pdf,
    },
    {
        'nombre'  : 'Banco — CSV/Excel con encabezados estándar',
        'tipo'    : 'BANCO',
        'ext'     : ['.csv', '.xlsx', '.xls'],
        'detectar': _det_banco_csv_generico,
        'parsear' : _par_banco_csv_generico,
    },
    {
        'nombre'  : 'Banco — TXT líneas fecha/descripción/valor',
        'tipo'    : 'BANCO',
        'ext'     : ['.txt'],
        'detectar': _det_banco_txt,
        'parsear' : _par_banco_txt,
    },
    # ── AUXILIAR ──────────────────────────────────────────────────────────────
    {
        'nombre'  : 'SIIGO — CSV Auxiliares Plan de Cuentas',
        'tipo'    : 'AUXILIAR',
        'ext'     : ['.csv'],
        'detectar': _det_siigo_aux_csv,
        'parsear' : _par_siigo_aux_csv,
    },
    {
        'nombre'  : 'Helisa — CSV Libro Auxiliar',
        'tipo'    : 'AUXILIAR',
        'ext'     : ['.csv', '.xlsx', '.xls'],
        'detectar': _det_helisa_aux_csv,
        'parsear' : _par_helisa_aux_csv,
    },
    {
        'nombre'  : 'World Office — CSV Auxiliar',
        'tipo'    : 'AUXILIAR',
        'ext'     : ['.csv', '.xlsx', '.xls'],
        'detectar': _det_world_office_aux_csv,
        'parsear' : _par_world_office_aux_csv,
    },
    {
        'nombre'  : 'SIIGO — PDF Auxiliares Plan de Cuentas',
        'tipo'    : 'AUXILIAR',
        'ext'     : ['.pdf'],
        'detectar': _det_siigo_aux_pdf,
        'parsear' : _par_aux_pdf_generico,
    },
    {
        'nombre'  : 'Auxiliar Contable — PDF (CON/CE/NC)',
        'tipo'    : 'AUXILIAR',
        'ext'     : ['.pdf'],
        'detectar': _det_aux_pdf_generico,
        'parsear' : _par_aux_pdf_generico,
    },
    {
        'nombre'  : 'Auxiliar Contable — CSV/Excel genérico',
        'tipo'    : 'AUXILIAR',
        'ext'     : ['.csv', '.xlsx', '.xls'],
        'detectar': _det_aux_csv_generico,
        'parsear' : _par_aux_csv_generico,
    },
    {
        'nombre'  : 'Auxiliar Contable — TXT',
        'tipo'    : 'AUXILIAR',
        'ext'     : ['.txt'],
        'detectar': _det_aux_txt,
        'parsear' : _par_aux_txt,
    },
]

# ── Despachador ───────────────────────────────────────────────────────────────

def _despachar(ruta, tipo, ext, usar_ocr):
    """Elige el mejor formato registrado y lo parsea. Retorna (df, meta, nombre_formato, confianza)."""
    candidatos = [f for f in REGISTRO_FORMATOS if f['tipo'] == tipo and ext in f['ext']]
    if not candidatos:
        raise ValueError(f"Formato no soportado para tipo={tipo}, extensión={ext}")
    muestra = _muestra_texto(ruta, ext)
    puntuaciones = [(f, f['detectar'](ruta, muestra)) for f in candidatos]
    mejor, conf = max(puntuaciones, key=lambda x: x[1])
    df, meta = mejor['parsear'](ruta, usar_ocr)
    return df, meta, mejor['nombre'], round(conf * 100)

# ── Función unificada de carga ────────────────────────────────────────────────
def cargar_y_parsear(uploaded_file, tipo, usar_ocr=False):
    nombre = uploaded_file.name
    ext = Path(nombre).suffix.lower()
    with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
        tmp.write(uploaded_file.getvalue())
        ruta = tmp.name

    try:
        if ext == '.pdf':
            diag = diagnosticar_pdf(ruta, tipo)
            ocr_efectivo = usar_ocr or diag['ocr_usado']
            df, res, fmt_nombre, fmt_conf = _despachar(ruta, tipo, ext, ocr_efectivo)
            legibilidad = (diag['pct_estimado_datos'], diag['calidad'],
                           diag['advertencias'], fmt_nombre, fmt_conf)
        elif ext in ['.csv', '.xlsx', '.xls']:
            df, res, fmt_nombre, fmt_conf = _despachar(ruta, tipo, ext, usar_ocr)
            legibilidad = (100.0, '🟢 EXCELENTE', [], fmt_nombre, fmt_conf)
        elif ext == '.txt':
            df, res, fmt_nombre, fmt_conf = _despachar(ruta, tipo, ext, usar_ocr)
            calidad = '🟢 EXCELENTE' if not df.empty else '🟠 PARCIAL'
            pct     = 95.0 if not df.empty else 50.0
            adv     = [] if not df.empty else ['Archivo TXT sin datos reconocibles']
            legibilidad = (pct, calidad, adv, fmt_nombre, fmt_conf)
        else:
            raise ValueError(f"Formato no soportado: {ext}")
    except Exception as e:
        raise e
    finally:
        os.unlink(ruta)
    return df, res, legibilidad

# ── Comparación inteligente (Fases A + B) ─────────────────────────────────
TOL_EXACTA = 1.0
TOL_APROX  = 0.005

# ── Fase A: clasificador de prefijo de documento auxiliar ─────────────────
_PAT_PREFIJO = re.compile(r'^([A-Z]{2,3})-', re.I)

def _prefijo_doc(doc_str):
    """Extrae prefijo: CE / CG / NC / CON / desconocido."""
    m = _PAT_PREFIJO.match(str(doc_str or ''))
    return m.group(1).upper() if m else ''

def _score_concepto(desc_banco, concepto_aux):
    """
    Similitud rápida entre descripción bancaria y concepto auxiliar.
    Devuelve 0.0–1.0 basado en palabras comunes (sin stopwords).
    """
    STOP = {'de','la','el','en','a','y','con','por','para','del','un','una',
            'los','las','al','se','su','que','no','es','pago','transferencia'}
    def _tokens(s):
        return {w.lower() for w in re.findall(r'[a-z0-9]{3,}', (s or '').lower())
                if w.lower() not in STOP}
    t1 = _tokens(desc_banco)
    t2 = _tokens(concepto_aux)
    if not t1 or not t2:
        return 0.0
    return len(t1 & t2) / max(len(t1), len(t2))

# ── Fase B: extractor de número de documento ─────────────────────────────
_PAT_NUMERICO = re.compile(r'[A-Z]{2,3}-(\d+)', re.I)

def _num_doc(doc_str):
    """Extrae la parte numérica de CE-250201 → '250201'."""
    m = _PAT_NUMERICO.match(str(doc_str or ''))
    return m.group(1) if m else ''

def comparar_documentos(df_b, df_a):
    """
    Reconciliación con matching inteligente por tipo de documento.

    Fase A — Restricción por prefijo:
        Movimiento ABONO  (vb > 0) → solo candidatos CG- / CON- (débitos)
        Movimiento CARGO  (vb < 0) → solo candidatos CE- / NC-   (créditos)
        Sin prefijo conocido        → candidatos libres de ambas columnas

    Fase B — Bonus por número de documento:
        Si el número del doc auxiliar (250201 de CE-250201) aparece
        en la descripción bancaria → ese candidato sube en prioridad.

    Dentro de los candidatos filtrados:
        1) Match exacto por monto + bonus de doc-num / similitud de concepto
        2) Match aproximado (±0.5 %) como fallback
        3) Sin match → '❌ SOLO EN BANCO'
    """
    if df_b.empty or df_a.empty:
        return pd.DataFrame(), df_a.copy() if not df_a.empty else pd.DataFrame()

    # ── Pre-cómputo ÚNICO antes del loop principal ───────────────────────────
    df_a = df_a.copy()
    df_a['_PREFIJO']  = df_a['DOCUMENTO'].apply(_prefijo_doc)
    df_a['_NUMERICO'] = df_a['DOCUMENTO'].apply(_num_doc)

    # Tokenizar conceptos auxiliares UNA SOLA VEZ (evita 1M+ re.findall)
    _STOP_SIM = {'de','la','el','en','a','y','con','por','para','del','un','una',
                 'los','las','al','se','su','que','no','es','pago','transferencia'}
    def _tok(s):
        return frozenset(
            w for w in re.findall(r'[a-z0-9]{3,}', (s or '').lower())
            if w not in _STOP_SIM
        )
    df_a['_CONC_TOK'] = df_a['CONCEPTO'].fillna('').apply(_tok)

    # Catálogo NC: cargar en memoria UNA vez
    _catalogo_nc_cache = []
    try:
        if os.path.exists(DB_PATH):
            _cn = sqlite3.connect(DB_PATH)
            _catalogo_nc_cache = _cn.execute(
                "SELECT banco_tokens, aux_tokens FROM nc_catalogo "
                "WHERE nivel IN ('ALTA','MEDIA') LIMIT 200"
            ).fetchall()
            _cn.close()
    except Exception:
        _catalogo_nc_cache = []

    # Pre-parsear tokens del catálogo NC (evita json.loads en el loop)
    _cat_parsed = []
    for _bt_j, _at_j in _catalogo_nc_cache:
        try:
            _cat_parsed.append((
                frozenset(json.loads(_bt_j or '[]')),
                frozenset(json.loads(_at_j or '[]'))
            ))
        except Exception:
            pass

    def _jaccard(a, b):
        if not a or not b: return 0.0
        return len(a & b) / len(a | b)

    idx_usados = set()
    filas      = []

    for idx_b, row_b in df_b.iterrows():
        vb = row_b['VALOR']
        if pd.isna(vb):
            continue

        monto_abs  = abs(vb)
        desc_banco = str(row_b.get('DESCRIPCION', '') or '')

        # Tokens de la descripción bancaria (calculados UNA vez por fila banco)
        banco_tok = _tok(desc_banco)

        # ── Fase A: filtrar candidatos por tipo ───────────────────────────
        es_abono = vb >= 0
        libres   = df_a[~df_a.index.isin(idx_usados)]

        if es_abono:
            col_buscar = 'DEBITO'
            candidatos = libres[
                libres['_PREFIJO'].isin(['CG','CON']) & libres[col_buscar].notna()
            ]
            if candidatos.empty:
                candidatos = libres[libres[col_buscar].notna()]
        else:
            col_buscar = 'CREDITO'
            candidatos = libres[
                libres['_PREFIJO'].isin(['CE','NC']) & libres[col_buscar].notna()
            ]
            if candidatos.empty:
                candidatos = libres[libres[col_buscar].notna()]

        candidatos = candidatos.copy()

        match_tipo = match_monto = match_idx = None
        match_doc  = match_conc = match_fecha_aux = ''
        match_metodo  = ''
        match_sim_val = 0.0

        if not candidatos.empty:
            candidatos['_diff'] = (candidatos[col_buscar] - monto_abs).abs()

            # ── Fase B: doc-num bonus (vectorizado) ───────────────────────
            candidatos['_doc_bonus'] = candidatos['_NUMERICO'].apply(
                lambda n: 1 if n and n in desc_banco else 0
            )

            # ── Similitud concepto con tokens PRE-computados ──────────────
            candidatos['_sim'] = candidatos['_CONC_TOK'].apply(
                lambda t: _jaccard(banco_tok, t)
            )

            # ── Bonus de proximidad de fecha (±5 días → +0.15) ───────────
            _fecha_b = row_b.get('FECHA')
            if pd.notna(_fecha_b) and 'FECHA' in candidatos.columns:
                candidatos['_fecha_bonus'] = candidatos['FECHA'].apply(
                    lambda f: 0.15 if (pd.notna(f) and
                        abs((pd.Timestamp(f) - pd.Timestamp(_fecha_b)).days) <= 10)
                    else 0.0
                )
            else:
                candidatos['_fecha_bonus'] = 0.0

            # ── Fase D: catálogo NC (solo si hay reglas y solo NC-) ───────
            candidatos['_cat_sim'] = 0.0
            if _cat_parsed:
                _nc_mask = candidatos['_PREFIJO'] == 'NC'
                if _nc_mask.any():
                    def _nc_cat_sim(aux_tok):
                        mejor = 0.0
                        for bt, at in _cat_parsed:
                            s = (_jaccard(banco_tok, bt) + _jaccard(aux_tok, at)) / 2
                            if s > mejor:
                                mejor = s
                        return mejor
                    candidatos.loc[_nc_mask, '_cat_sim'] = \
                        candidatos.loc[_nc_mask, '_CONC_TOK'].apply(_nc_cat_sim)

            # ── Score combinado ───────────────────────────────────────────────
            exactos = candidatos[candidatos['_diff'] <= TOL_EXACTA].copy()
            if not exactos.empty:
                exactos = exactos.sort_values(
                    ['_doc_bonus', '_cat_sim', '_fecha_bonus', '_sim', '_diff'],
                    ascending=[False, False, False, False, True]
                )
                mejor = exactos.iloc[0]
                match_tipo    = 'EXACTO'
                if mejor['_doc_bonus']:
                    match_metodo = 'DOC+MONTO'
                elif mejor['_cat_sim'] >= 0.30:
                    match_metodo = 'CATALOGO+MONTO'
                else:
                    match_metodo = 'MONTO'
                match_monto     = mejor[col_buscar]
                match_idx       = mejor.name
                match_doc       = mejor['DOCUMENTO']
                match_conc      = mejor['CONCEPTO']
                match_fecha_aux = mejor['FECHA_RAW']
                match_sim_val   = float(mejor.get('_sim', 0.0))

            # ── Fallback: match aproximado ────────────────────────────────────
            if match_tipo is None and monto_abs > 0:
                aprox = candidatos[
                    (candidatos['_diff'] / monto_abs) <= TOL_APROX
                ].copy()
                if not aprox.empty:
                    aprox = aprox.sort_values(
                        ['_doc_bonus', '_cat_sim', '_fecha_bonus', '_sim', '_diff'],
                        ascending=[False, False, False, False, True]
                    )
                    mejor = aprox.iloc[0]
                    match_tipo    = 'APROX'
                    if mejor['_doc_bonus']:
                        match_metodo = 'DOC+APROX'
                    elif mejor['_cat_sim'] >= 0.30:
                        match_metodo = 'CATALOGO+APROX'
                    else:
                        match_metodo = 'APROX'
                    match_monto     = mejor[col_buscar]
                    match_idx       = mejor.name
                    match_doc       = mejor['DOCUMENTO']
                    match_conc      = mejor['CONCEPTO']
                    match_fecha_aux = mejor['FECHA_RAW']
                    match_sim_val   = float(mejor.get('_sim', 0.0))

        if match_idx is not None:
            idx_usados.add(match_idx)

        estado   = ('✅ COINCIDE EXACTO' if match_tipo == 'EXACTO'
                    else '🔶 COINCIDE APROX.' if match_tipo == 'APROX'
                    else '❌ SOLO EN BANCO')
        diff_val = abs(monto_abs - match_monto) if match_monto is not None else None

        # Calcular confianza del match
        _confianza = {
            'DOC+MONTO'      : 95,
            'CATALOGO+MONTO' : 85,
            'MONTO'          : max(60, 60 + int(match_sim_val * 25)),
            'DOC+APROX'      : 75,
            'CATALOGO+APROX' : 60,
            'APROX'          : max(40, 40 + int(match_sim_val * 20)),
        }.get(match_metodo, 0)

        filas.append({
            'N'              : idx_b,
            'FECHA_BANCO'    : row_b['FECHA_RAW'],
            'TIPO_MOV'       : row_b['TIPO'],
            'DESCRIPCION'    : desc_banco,
            'VALOR_BANCO'    : vb,
            'DOC_AUXILIAR'   : match_doc,
            'FECHA_AUXILIAR' : match_fecha_aux,
            'CONCEPTO_AUX'   : match_conc,
            'MONTO_AUXILIAR' : match_monto,
            'DIFERENCIA'     : diff_val,
            'ESTADO'         : estado,
            'MATCH_TIPO'     : match_tipo or 'SIN_MATCH',
            'METODO_MATCH'   : match_metodo,
            'CONFIANZA'      : _confianza,
            'PAGINA_PDF'     : row_b.get('PAGINA', ''),
        })

    df_comp = pd.DataFrame(filas)
    df_solo_aux = df_a[~df_a.index.isin(idx_usados)].copy()
    # Limpiar columnas internas del auxiliar
    for _c in ['_PREFIJO', '_NUMERICO', '_CONC_TOK']:
        if _c in df_solo_aux.columns:
            df_solo_aux.drop(columns=[_c], inplace=True)
    df_solo_aux['ESTADO'] = '📋 SOLO EN AUXILIAR'

    # ══════════════════════════════════════════════════════════════════════
    # FASE E — Segundo paso: cargos rechazados sin asiento (tolerancia ±3%)
    # Detecta cargos bancarios con keywords de rechazo/devolución y los
    # empareja con NC- del auxiliar que quedaron sin match en el loop principal.
    # Se muestra como '🔄 RECHAZO — CONFIRMAR' para revisión humana.
    # ══════════════════════════════════════════════════════════════════════
    _PAT_REC_B = re.compile(
        r'RECHAZOS?|DEBITO\s+POR|ND\s+POR|DEVOLUCI|ANULACI|RETORNO|REVERSO|COBRO\s+INV|REINTEGRO', re.I)
    _PAT_REC_A = re.compile(
        r'RECHAZOS?|DEVOLUCI|ANULACI|RETORNO|REVERSO|REINTEGRO|NOTA\s+CONT|COMISI|PAGOS\s+A', re.I)
    _TOL_RECHAZO = 0.05   # ±5 % — cubre diferencias de comisión (ej. 2682.17 vs 2659.92 = 0.84%)

    if not df_comp.empty and not df_solo_aux.empty:
        # NC- libres en el auxiliar (solo las que quedaron sin emparejar)
        _nc_libres = df_solo_aux[
            df_solo_aux.get('DOCUMENTO', pd.Series(dtype=str))
                        .str.startswith('NC-', na=False) &
            df_solo_aux['CREDITO'].notna()
        ].copy()

        _usados_fase_e = set()

        for _fi in df_comp[df_comp['ESTADO'] == '❌ SOLO EN BANCO'].index:
            _rc = df_comp.loc[_fi]
            if str(_rc.get('TIPO_MOV', '') or '') != 'CARGO':
                continue
            _desc_b = str(_rc.get('DESCRIPCION', '') or '')
            if not _PAT_REC_B.search(_desc_b):
                continue
            _monto_b = abs(float(_rc.get('VALOR_BANCO', 0) or 0))
            if _monto_b < 1:
                continue

            # NC libres con diferencia de monto dentro de la tolerancia
            _cands = _nc_libres[
                ~_nc_libres.index.isin(_usados_fase_e) &
                ((_nc_libres['CREDITO'] - _monto_b).abs() / _monto_b <= _TOL_RECHAZO)
            ].copy()
            if _cands.empty:
                continue

            # Priorizar NC que también tengan keywords de rechazo en su concepto
            _cands['_rb'] = _cands['CONCEPTO'].fillna('').apply(
                lambda _c: 2 if _PAT_REC_A.search(_c) else 0)
            _cands['_dr'] = (_cands['CREDITO'] - _monto_b).abs()
            _mejor_r = _cands.sort_values(['_rb', '_dr'],
                                          ascending=[False, True]).iloc[0]

            df_comp.loc[_fi, 'DOC_AUXILIAR']  = _mejor_r.get('DOCUMENTO', '')
            df_comp.loc[_fi, 'FECHA_AUXILIAR'] = _mejor_r.get('FECHA_RAW', '')
            df_comp.loc[_fi, 'CONCEPTO_AUX']  = _mejor_r.get('CONCEPTO', '')
            df_comp.loc[_fi, 'MONTO_AUXILIAR'] = _mejor_r['CREDITO']
            df_comp.loc[_fi, 'DIFERENCIA']    = abs(_monto_b - _mejor_r['CREDITO'])
            df_comp.loc[_fi, 'ESTADO']        = '🔄 RECHAZO — CONFIRMAR'
            df_comp.loc[_fi, 'MATCH_TIPO']    = 'RECHAZO'
            df_comp.loc[_fi, 'METODO_MATCH']  = 'FASE_E'
            df_comp.loc[_fi, 'CONFIANZA']     = 45

            _usados_fase_e.add(_mejor_r.name)

        # Quitar del auxiliar suelto las NC que Fase E emparejó
        if _usados_fase_e:
            df_solo_aux = df_solo_aux.drop(index=list(_usados_fase_e), errors='ignore')

    # ══════════════════════════════════════════════════════════════════════
    # FASE F — N cargos bancarios → 1 NC (matching por agrupación)
    # Cuando el banco cobra N veces el mismo tipo de cargo (ej: IVA por
    # cada transacción) y el auxiliar tiene UNA sola NC por el total.
    # Tolerancia ±1 % para cubrir redondeos del contador.
    # ══════════════════════════════════════════════════════════════════════
    _TOL_GRUPO = 0.01   # ±1 %

    if not df_comp.empty and not df_solo_aux.empty:
        # Cargos bancarios que siguen SOLO EN BANCO
        _sb_f = df_comp[
            (df_comp['ESTADO'] == '❌ SOLO EN BANCO') &
            (df_comp['TIPO_MOV'] == 'CARGO')
        ].copy()

        # NC libres en el auxiliar
        _nc_f = df_solo_aux[
            df_solo_aux.get('DOCUMENTO', pd.Series(dtype=str))
                        .str.startswith('NC-', na=False) &
            df_solo_aux['CREDITO'].notna()
        ].copy()

        if not _sb_f.empty and not _nc_f.empty:
            # Clave de agrupación: tokens significativos de la descripción
            def _clave_grupo(s):
                words = re.findall(r'[A-Z]{3,}', (s or '').upper())
                # quitar palabras muy genéricas
                _skip = {'CARG','CARGO','PAGO','PROV','BANC','COBR'}
                return '|'.join(w for w in words if w not in _skip)

            _sb_f['_gkey'] = _sb_f['DESCRIPCION'].apply(_clave_grupo)

            _usados_f_b  = set()
            _usados_f_nc = set()

            # Para cada NC libre, buscar un grupo cuya suma coincida
            for _nci, _ncrow in _nc_f.iterrows():
                if _nci in _usados_f_nc:
                    continue
                _nc_val = float(_ncrow['CREDITO'])
                if _nc_val < 1:
                    continue

                _pendientes = _sb_f[~_sb_f.index.isin(_usados_f_b)]
                if _pendientes.empty:
                    break

                _mejor_grupo_idx  = None
                _mejor_grupo_diff = None

                for _gkey, _grp in _pendientes.groupby('_gkey'):
                    if not _gkey:
                        continue
                    _suma = float(_grp['VALOR_BANCO'].abs().sum())
                    if _suma < 1:
                        continue
                    _diff_pct = abs(_suma - _nc_val) / max(_nc_val, 1)
                    if _diff_pct <= _TOL_GRUPO:
                        # Preferir el grupo cuya suma sea más cercana
                        if _mejor_grupo_diff is None or _diff_pct < _mejor_grupo_diff:
                            _mejor_grupo_idx  = list(_grp.index)
                            _mejor_grupo_diff = _diff_pct

                if _mejor_grupo_idx is None:
                    continue

                _n = len(_mejor_grupo_idx)
                _conf_f = max(55, int((1 - _mejor_grupo_diff) * 85))
                for _bidx in _mejor_grupo_idx:
                    df_comp.loc[_bidx, 'DOC_AUXILIAR']  = _ncrow.get('DOCUMENTO', '')
                    df_comp.loc[_bidx, 'FECHA_AUXILIAR'] = _ncrow.get('FECHA_RAW', '')
                    df_comp.loc[_bidx, 'CONCEPTO_AUX']  = _ncrow.get('CONCEPTO', '')
                    df_comp.loc[_bidx, 'MONTO_AUXILIAR'] = round(_nc_val / _n, 2)
                    df_comp.loc[_bidx, 'DIFERENCIA']    = round(
                        abs(float(df_comp.loc[_bidx,'VALOR_BANCO']) + _nc_val/_n), 2)
                    df_comp.loc[_bidx, 'ESTADO']        = f'🔵 AGRUPADO N:1 ({_n} cargos → 1 NC)'
                    df_comp.loc[_bidx, 'MATCH_TIPO']    = 'AGRUPADO'
                    df_comp.loc[_bidx, 'METODO_MATCH']  = f'FASE_F_N{_n}'
                    df_comp.loc[_bidx, 'CONFIANZA']     = _conf_f
                    _usados_f_b.add(_bidx)
                _usados_f_nc.add(_nci)

        # Quitar NC usadas en Fase F del auxiliar suelto
        if not _sb_f.empty and '_usados_f_nc' in dir() and _usados_f_nc:
            df_solo_aux = df_solo_aux.drop(index=list(_usados_f_nc), errors='ignore')

    return df_comp, df_solo_aux

# ══════════════════════════════════════════════════════════════════════════════
# INTERFAZ PREMIUM — solo capa de presentación, sin tocar lógica de datos
# ══════════════════════════════════════════════════════════════════════════════

# ── CSS Global ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
/* ════════════════════════════════════════════════
   PREMIUM MINIMALIST — funciona en tema oscuro y claro
   Estrategia: rgba semi-transparente + texto heredado
   ════════════════════════════════════════════════ */

/* ── Fuente global ── */
html, body, [class*="css"] {
  font-family: 'Segoe UI', system-ui, -apple-system, sans-serif;
}

/* ── Tarjetas métricas — borde izquierdo azul, fondo neutro semi-transparente ── */
[data-testid="metric-container"] {
    background: rgba(66, 165, 245, 0.08) !important;
    border: 1px solid rgba(66, 165, 245, 0.25) !important;
    border-left: 4px solid #42a5f5 !important;
    border-radius: 10px;
    padding: 16px 20px;
}
[data-testid="metric-container"] label {
    font-size: 0.75rem;
    font-weight: 700;
    letter-spacing: .07em;
    text-transform: uppercase;
    opacity: 0.75;
}
[data-testid="metric-container"] [data-testid="stMetricValue"] {
    color: #42a5f5 !important;
    font-size: 1.3rem;
    font-weight: 800;
}
[data-testid="metric-container"] [data-testid="stMetricDelta"] {
    font-size: 0.78rem;
    font-weight: 700;
}

/* ── Tabs ── */
button[data-baseweb="tab"] {
    font-weight: 600;
    font-size: 0.85rem;
}
button[data-baseweb="tab"][aria-selected="true"] {
    color: #42a5f5 !important;
    border-bottom: 3px solid #42a5f5 !important;
}

/* ── Dataframes ── */
[data-testid="stDataFrame"] {
    border-radius: 10px;
    overflow: hidden;
    border: 1px solid rgba(66,165,245,0.15);
}

/* ── Sidebar — gradiente oscuro, texto blanco ── */
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #0d1b2e 0%, #0a1628 100%) !important;
    border-right: 1px solid rgba(66,165,245,0.2);
}
[data-testid="stSidebar"] * { color: #e0e8f4 !important; }
[data-testid="stSidebar"] .stButton > button {
    background: rgba(66,165,245,0.15) !important;
    border: 1px solid rgba(66,165,245,0.4) !important;
    border-radius: 8px;
    font-weight: 700;
    color: #ffffff !important;
    transition: background 0.2s;
}
[data-testid="stSidebar"] .stButton > button:hover {
    background: rgba(66,165,245,0.30) !important;
}

/* ── Expanders ── */
details > summary {
    font-weight: 700;
    font-size: 0.92rem;
}

/* ══ CALLOUT BOXES — rgba + borde izquierdo coloreado ══
   Texto: color:inherit  →  adapta a tema oscuro/claro      */

.callout-info {
    background: rgba(66, 165, 245, 0.10);
    border-left: 4px solid #42a5f5;
    border-radius: 8px;
    padding: 13px 17px;
    margin: 9px 0;
    color: inherit;
    line-height: 1.6;
}
.callout-success {
    background: rgba(102, 187, 106, 0.12);
    border-left: 4px solid #66bb6a;
    border-radius: 8px;
    padding: 13px 17px;
    margin: 9px 0;
    color: inherit;
    line-height: 1.6;
}
.callout-warning {
    background: rgba(255, 183, 77, 0.12);
    border-left: 4px solid #ffb74d;
    border-radius: 8px;
    padding: 13px 17px;
    margin: 9px 0;
    color: inherit;
    line-height: 1.6;
}
.callout-danger {
    background: rgba(239, 83, 80, 0.12);
    border-left: 4px solid #ef5350;
    border-radius: 8px;
    padding: 13px 17px;
    margin: 9px 0;
    color: inherit;
    line-height: 1.6;
}
.callout-accion {
    background: rgba(171, 71, 188, 0.12);
    border-left: 4px solid #ab47bc;
    border-radius: 8px;
    padding: 13px 17px;
    margin: 9px 0;
    color: inherit;
    font-family: 'Consolas', 'Courier New', monospace;
    font-size: 0.88rem;
    line-height: 1.6;
}

/* ══ BADGES — texto siempre oscuro para máximo contraste ══ */
.badge-verde  {
    display: inline-block;
    background: #1b5e20;
    color: #a5d6a7 !important;
    border-radius: 20px;
    padding: 3px 11px;
    font-size: .78rem;
    font-weight: 800;
}
.badge-rojo   {
    display: inline-block;
    background: #b71c1c;
    color: #ffcdd2 !important;
    border-radius: 20px;
    padding: 3px 11px;
    font-size: .78rem;
    font-weight: 800;
}
.badge-naranja {
    display: inline-block;
    background: #e65100;
    color: #ffe0b2 !important;
    border-radius: 20px;
    padding: 3px 11px;
    font-size: .78rem;
    font-weight: 800;
}
.badge-azul   {
    display: inline-block;
    background: #0d47a1;
    color: #bbdefb !important;
    border-radius: 20px;
    padding: 3px 11px;
    font-size: .78rem;
    font-weight: 800;
}

/* ══ HEADER PRINCIPAL ══ */
.main-header {
    background: linear-gradient(135deg, #0d47a1 0%, #1565C0 60%, #1e88e5 100%);
    color: #ffffff;
    padding: 26px 34px;
    border-radius: 14px;
    margin-bottom: 22px;
    border: 1px solid rgba(66,165,245,0.3);
    box-shadow: 0 4px 24px rgba(13,71,161,0.35);
}
.main-header h1 {
    color: #ffffff !important;
    margin: 0;
    font-size: 1.65rem;
    font-weight: 800;
    letter-spacing: -.01em;
}
.main-header p {
    color: rgba(255,255,255,0.72) !important;
    margin: 6px 0 0 0;
    font-size: .92rem;
}

/* ══ TÍTULOS DE SECCIÓN ══ */
.section-title {
    color: #42a5f5;
    font-size: 1.0rem;
    font-weight: 800;
    letter-spacing: .04em;
    text-transform: uppercase;
    border-bottom: 1px solid rgba(66,165,245,0.3);
    padding-bottom: 7px;
    margin: 18px 0 12px 0;
}

/* ══ FILAS DE GUÍA ══ */
.guia-row {
    background: rgba(66, 165, 245, 0.07);
    border-radius: 8px;
    padding: 13px 17px;
    margin: 7px 0;
    border: 1px solid rgba(66,165,245,0.20);
    line-height: 1.7;
    transition: background 0.15s, border-color 0.15s;
}
.guia-row:hover {
    background: rgba(66, 165, 245, 0.14);
    border-color: rgba(66,165,245,0.45);
}

/* ══ PROGRESS BAR ══ */
[data-testid="stProgressBar"] > div > div {
    background: linear-gradient(90deg, #1565C0, #42a5f5) !important;
    border-radius: 4px;
}
</style>
""", unsafe_allow_html=True)

# ── Control de acceso ─────────────────────────────────────────────────────────
_PWD_OK = os.environ.get("APP_PASSWORD", "crediexpress2025")

if "autenticado" not in st.session_state:
    st.session_state.autenticado = False

if not st.session_state.autenticado:
    st.markdown("""
    <div style='max-width:420px;margin:80px auto 0 auto;'>
      <div class='main-header' style='text-align:center;padding:32px 40px;'>
        <div style='font-size:2.5rem;margin-bottom:8px;'>🔒</div>
        <h1 style='font-size:1.4rem;'>Acceso Restringido</h1>
        <p>CREDIEXPRESS POPAYÁN SAS</p>
      </div>
    </div>""", unsafe_allow_html=True)
    col_l, col_c, col_r = st.columns([1, 2, 1])
    with col_c:
        pwd = st.text_input("Contraseña", type="password", placeholder="Ingrese la contraseña...",
                            label_visibility="collapsed")
        if st.button("Ingresar", use_container_width=True):
            if pwd == _PWD_OK:
                st.session_state.autenticado = True
                st.rerun()
            else:
                st.error("Contraseña incorrecta. Intente de nuevo.")
        st.caption("💡 Contacte al administrador si olvidó la contraseña.")
    st.stop()

# ── Helpers de análisis (solo UI, no tocan datos) ────────────────────────────

def _cop_limpio(v):
    """Versión limpia del cop() para HTML."""
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return 'N/A'
    return f"${abs(v):,.0f}" + (" CR" if v < 0 else "")

def _semaforo_conciliacion(pct):
    if pct >= 90: return "🟢", "EXCELENTE", "verde"
    if pct >= 75: return "🟡", "BUENA",     "naranja"
    if pct >= 50: return "🟠", "REGULAR",   "naranja"
    return "🔴", "CRÍTICA", "rojo"

def _inferir_cuenta_sugerida(desc, valor):
    """Sugiere cuenta contable para movimientos bancarios sin asiento."""
    d = (desc or '').upper()
    if any(x in d for x in ['GMF','4X1000','IMPTO GOBIERNO']):
        return '5305 — Impuesto GMF 4×1000', 'NC'
    if any(x in d for x in ['COMISION','COMISIÓN']):
        return '5305 — Comisiones Bancarias', 'NC'
    if 'NEQUI' in d:
        return '5305 — Comisiones Nequi/PSE', 'NC'
    if 'PSE' in d and valor < 0:
        return '5305 — Comisiones PSE', 'NC'
    if any(x in d for x in ['INTERES','INTERÉS','RENDIMIENTO']):
        return ('4205 — Rendimientos Financieros', 'CE') if valor > 0 else ('5305 — Intereses Débito', 'NC')
    if any(x in d for x in ['PAGO A PROV','PAGO A PROVE','PAGO PROVE']):
        return '2205 — Proveedores', 'CE'
    if any(x in d for x in ['NOMINA','NÓMINA','SALARIO']):
        return '2335 — Nómina por Pagar', 'CE'
    if any(x in d for x in ['TRANSFERENCIA','TRASLADO']):
        return '1110 — Bancos (verificar destino)', 'CE' if valor > 0 else 'NC'
    if valor > 0:
        return '1305 — Clientes / Recaudo (verificar)', 'CE'
    return '5999 — Otros Gastos (verificar)', 'NC'

def _guia_banco_sin_aux(row):
    """Genera instrucción específica para movimiento bancario sin asiento."""
    fecha  = row.get('FECHA_BANCO', '')
    desc   = str(row.get('DESCRIPCION', ''))[:60]
    valor  = row.get('VALOR_BANCO', 0)
    tipo   = row.get('TIPO_MOV', '')
    pagina = row.get('PAGINA_PDF', '')
    ref_pagina = f"Página <b>{pagina}</b> del PDF" if pagina else "Ver extracto bancario"
    cuenta, comprobante = _inferir_cuenta_sugerida(desc, valor)
    signo  = "+" if valor > 0 else "-"
    return f"""
<div class='guia-row'>
<b>📍 UBICAR EN EXTRACTO BANCARIO</b><br>
&nbsp;&nbsp;Fecha: <b>{fecha}</b> &nbsp;|&nbsp; Tipo: <b>{tipo}</b> &nbsp;|&nbsp; Valor: <b>{signo}${abs(valor):,.0f}</b> &nbsp;|&nbsp; {ref_pagina}<br>
&nbsp;&nbsp;Descripción: <i>"{desc}"</i><br><br>
<b>✏️ ACCIÓN EN SISTEMA CONTABLE</b><br>
&nbsp;&nbsp;Crear comprobante: <b>{comprobante}-XXXXXX</b><br>
&nbsp;&nbsp;Fecha: <b>{fecha}</b> &nbsp;|&nbsp; Cuenta sugerida: <b>{cuenta}</b><br>
&nbsp;&nbsp;Valor: <b>${abs(valor):,.0f}</b>
</div>"""

def _guia_aux_sin_banco(row):
    """Genera instrucción específica para asiento contable sin transacción bancaria."""
    doc     = str(row.get('DOCUMENTO', ''))
    fecha   = str(row.get('FECHA_RAW', ''))
    concepto= str(row.get('CONCEPTO', ''))[:60]
    deb     = row.get('DEBITO',  None)
    cre     = row.get('CREDITO', None)
    valor   = deb if deb else cre
    col     = row.get('COLUMNA', '')
    return f"""
<div class='guia-row'>
<b>📋 DOCUMENTO EN AUXILIAR CONTABLE</b><br>
&nbsp;&nbsp;Documento: <b>{doc}</b> &nbsp;|&nbsp; Fecha: <b>{fecha}</b> &nbsp;|&nbsp; Tipo: <b>{col}</b><br>
&nbsp;&nbsp;Concepto: <i>"{concepto}"</i> &nbsp;|&nbsp; Valor: <b>${abs(valor or 0):,.0f}</b><br><br>
<b>🔍 BUSCAR EN EXTRACTO BANCARIO</b><br>
&nbsp;&nbsp;Buscar movimiento de <b>${abs(valor or 0):,.0f} COP</b> cerca del <b>{fecha}</b><br>
&nbsp;&nbsp;Si no aparece: verificar si fue anulado, está en otro período o es asiento de ajuste interno.
</div>"""

def _extraer_periodo(nombre_archivo):
    """Extrae (anio, mes) del nombre del archivo. Ej: FEBRERO_2025 -> (2025, 2)."""
    meses = {
        'enero':1,'febrero':2,'marzo':3,'abril':4,'mayo':5,'junio':6,
        'julio':7,'agosto':8,'septiembre':9,'octubre':10,'noviembre':11,'diciembre':12,
        'january':1,'february':2,'march':3,'april':4,'may':5,'june':6,
        'july':7,'august':8,'september':9,'october':10,'november':11,'december':12,
    }
    n = (nombre_archivo or '').lower()
    anio_m = re.search(r'(20\d{2})', n)
    mes_num = next((v for k, v in meses.items() if k in n), None)
    if anio_m and mes_num:
        return (int(anio_m.group(1)), mes_num)
    return None

_MESES_ES = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
             'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']

# ── Header dinámico ───────────────────────────────────────────────────────────
st.markdown("""
<div class='main-header'>
  <h1>🏦 Conciliación Bancaria — CREDIEXPRESS / TRASNODUS SAS</h1>
  <p>Extracto Bancolombia &nbsp;↔&nbsp; Auxiliar Contable &nbsp;·&nbsp;
     Detección automática de formato &nbsp;·&nbsp; Análisis inteligente de diferencias</p>
</div>
""", unsafe_allow_html=True)

with st.sidebar:
    st.markdown("""
    <div style='text-align:center;padding:8px 0 16px 0;'>
      <div style='font-size:2rem;'>🏦</div>
      <div style='font-size:1.1rem;font-weight:800;letter-spacing:.03em;'>CREDIEXPRESS</div>
      <div style='font-size:.78rem;opacity:.85;'>Conciliación Bancaria · Sistema Inteligente</div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("#### 📂 Cargar Archivos")
    banco_file = st.file_uploader("Extracto Bancolombia", type=["pdf","csv","xlsx","txt"],
                                   help="PDF original del banco, CSV con columnas Fecha/Descripción/Valor/Saldo, o Excel.")
    aux_file   = st.file_uploader("Auxiliar Contable", type=["pdf","csv","xlsx","txt"],
                                   help="CSV / Excel exportado desde SIIGO, Helisa, World Office con columnas Documento/Fecha/Concepto/Debito/Credito.")

    usar_ocr = st.checkbox("🔍 Forzar OCR en PDF escaneados", value=True,
                            help="Requiere Tesseract + Poppler instalados.")

    if banco_file:
        st.markdown(f"<div style='background:#ffffff22;border-radius:6px;padding:6px 10px;font-size:.8rem;'>📄 <b>{banco_file.name}</b><br><span style='opacity:.75'>{banco_file.size/1024:.1f} KB</span></div>", unsafe_allow_html=True)
    if aux_file:
        st.markdown(f"<div style='background:#ffffff22;border-radius:6px;padding:6px 10px;font-size:.8rem;margin-top:4px;'>📄 <b>{aux_file.name}</b><br><span style='opacity:.75'>{aux_file.size/1024:.1f} KB</span></div>", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    ejecutar = st.button("🚀 Ejecutar análisis completo", disabled=not (banco_file and aux_file),
                          use_container_width=True)
    if ejecutar:
        st.session_state.run = True

    st.markdown("---")
    st.markdown("#### 💡 Formatos soportados")
    st.markdown("""
- **PDF** — extracto bancario original
- **CSV** — SIIGO / Helisa / World Office
- **Excel (.xlsx)** — cualquier formato tabular
- **TXT** — texto plano con columnas
    """)
    if OCR_AVAILABLE:
        st.markdown("<span class='badge-verde'>✅ OCR disponible</span>", unsafe_allow_html=True)
    else:
        st.markdown("<span class='badge-naranja'>⚠️ OCR no instalado</span>", unsafe_allow_html=True)
        st.caption("Instale pytesseract + Poppler para PDFs escaneados.")

    st.markdown("---")
    st.markdown("<div style='font-size:.72rem;opacity:.7;text-align:center;'>v2.0 · CREDIEXPRESS POPAYÁN SAS<br>Desarrollado con ❤️ en Python + Streamlit</div>", unsafe_allow_html=True)
    # ── Auto-guardar archivos subidos (solo offline) ─────────────────────────
    if OFFLINE_MODE:
        for _uf, _sub in [(banco_file, "datos_entrada"), (aux_file, "datos_entrada")]:
            if _uf:
                _ruta, _nuevo = _auto_guardar_archivo(_uf, _sub)
                if _nuevo and _ruta:
                    st.caption(f"💾 Guardado: .../{_uf.name}")

    # ── Panel historial (solo offline) ───────────────────────────────────────
    if OFFLINE_MODE:
        # ── Fase C: formatos aprendidos ──────────────────────────────────
        _formatos = listar_formatos_aprendidos()
        if _formatos:
            st.markdown("---")
            st.markdown("**🧠 Formatos Aprendidos**")
            for _fma, _tipo, _banco, _usos, _ultima in _formatos[:5]:
                _ico2 = "📄" if _tipo == "auxiliar" else "🏦"
                st.markdown(
                    f"<small>{_ico2} <b>{_fma}</b><br>"
                    f"&nbsp;&nbsp;{_usos} uso{'s' if _usos!=1 else ''}"
                    f" · {(_ultima or '')[:10]}</small>",
                    unsafe_allow_html=True
                )
        # ── Fase D5: catalogo NC aprendido ────────────────────────────────
        try:
            _cat_rows, _cat_total, _cat_pend = listar_catalogo_nc(5)
        except Exception:
            _cat_rows, _cat_total, _cat_pend = [], 0, 0
        st.markdown("---")
        st.markdown(
            f"**📚 Catálogo NC** &nbsp;"
            f"<span style='color:#42a5f5;font-size:.8rem;'>"
            f"{_cat_total} reglas · {_cat_pend} pendientes</span>",
            unsafe_allow_html=True
        )
        if _cat_rows:
            for _cr in _cat_rows:
                _uuid, _bt, _at, _conf, _nivel, _apr, _ult = _cr
                _ico_nv = ("🟢" if _nivel=='ALTA'
                           else "🟡" if _nivel=='MEDIA'
                           else "⏳")
                try:
                    _bt_tok = json.loads(_bt or '[]')[:3]
                    _at_tok = json.loads(_at or '[]')[:3]
                    _lbl = ' '.join(_bt_tok) + ' ↔ ' + ' '.join(_at_tok)
                except Exception:
                    _lbl = _uuid
                st.markdown(
                    f"<small>{_ico_nv} {_lbl}<br>"
                    f"&nbsp;&nbsp;{_conf} confirmaci{'o' if _conf==1 else 'o'}nes"
                    f" · {(_ult or '')[:10]}</small>",
                    unsafe_allow_html=True
                )
        else:
            st.markdown(
                "<small style='opacity:.5;'>Aun sin reglas aprendidas.<br>"
                "Se llenara automaticamente al procesar PDFs.</small>",
                unsafe_allow_html=True
            )
        # ── Boton de sincronizacion ───────────────────────────────────────
        st.markdown("")
        if st.button("🔄 Sincronizar con Cloud", use_container_width=True,
                     help="Sube reglas nuevas a Google Sheets y baja las del cloud"):
            with st.spinner("Sincronizando..."):
                _n_up, _n_down = sincronizar_catalogo_nc()
            st.success(
                f"✅ Sync OK — "
                f"↑{_n_up} subidas · ↓{_n_down} bajadas"
            )
        _hist = leer_historial_sqlite(6)
        if _hist:
            st.markdown("---")
            st.markdown("#### 📜 Historial")
            for _h in _hist:
                _fh, _fb, _fa, _per, _tasa, _ex, _nb, _dif = _h
                _ico = "🟢" if _tasa >= 90 else ("🟡" if _tasa >= 75 else "🔴")
                _per_lbl = f" · {_per}" if _per else ""
                st.markdown(f"""
<div style='background:rgba(255,255,255,0.08);border-radius:6px;
            padding:7px 10px;margin:3px 0;font-size:.74rem;line-height:1.5;'>
  {_ico} <b>{_fh[:16]}</b>{_per_lbl}<br>
  <span style='opacity:.65;'>{os.path.basename(_fb)[:26]}</span><br>
  <span style='color:#42a5f5;font-weight:700;'>{_tasa:.0f}% conciliado
    &nbsp;·&nbsp; {_ex}/{_nb} mov.</span>
</div>""", unsafe_allow_html=True)


if 'run' in st.session_state and st.session_state.run:
    with st.spinner("Procesando archivos..."):
        try:
            df_banco, res_banco, leg_banco = cargar_y_parsear(banco_file, 'BANCO', usar_ocr=usar_ocr)
            sa  = res_banco.get('SALDO_INICIAL', 0) or res_banco.get('SALDO_ANTERIOR', 0) or 0
            sac = res_banco.get('SALDO_FINAL', 0) or res_banco.get('SALDO_ACTUAL', 0) or 0
            tab_s = res_banco.get('TOTAL_ABONOS', 0) or 0
            tca_s = abs(res_banco.get('TOTAL_CARGOS', 0)) or 0  # cargos negativos en CSV los ponemos absolutos

            df_aux, meta_aux, leg_aux = cargar_y_parsear(aux_file, 'AUXILIAR', usar_ocr=usar_ocr)
            si_a = meta_aux.get('SALDO_INICIAL', 0) or 0
            sf_a = meta_aux.get('SALDO_FINAL',   0) or 0
            td_a = meta_aux.get('TOTAL_DEBITOS', 0) or 0
            tc_a = meta_aux.get('TOTAL_CREDITOS',0) or 0
        except Exception as e:
            st.error(f"❌ Error al procesar los archivos: {e}")
            st.stop()

    # ── Validación de período ─────────────────────────────────────────────────
    periodo_b = _extraer_periodo(banco_file.name)
    periodo_a = _extraer_periodo(aux_file.name)
    if periodo_b and periodo_a and periodo_b != periodo_a:
        st.markdown(f"""
        <div class='callout-danger'>
          <b>⚠️ ALERTA DE PERÍODO — Los archivos son de meses distintos</b><br>
          Extracto bancario: <b>{_MESES_ES[periodo_b[1]-1]} {periodo_b[0]}</b>
          &nbsp;·&nbsp;
          Auxiliar contable: <b>{_MESES_ES[periodo_a[1]-1]} {periodo_a[0]}</b><br>
          Para una conciliación correcta ambos archivos deben ser del mismo mes y año.
          Verifique los archivos cargados antes de continuar.
        </div>""", unsafe_allow_html=True)
    elif periodo_b and periodo_a:
        st.markdown(f"""
        <div class='callout-success' style='padding:8px 14px;font-size:.88rem;'>
          ✅ Período verificado: ambos archivos corresponden a
          <b>{_MESES_ES[periodo_b[1]-1]} {periodo_b[0]}</b>
        </div>""", unsafe_allow_html=True)

    # Comparación
    if not df_aux.empty:
        df_comp, df_solo_aux = comparar_documentos(df_banco, df_aux)
        n_tot    = len(df_comp)
        rechazos = df_comp[df_comp['ESTADO'] == '🔄 RECHAZO — CONFIRMAR']
        agrupados = df_comp[df_comp['ESTADO'].str.startswith('🔵 AGRUPADO', na=False)]
        # ── Fase D3: auto-aprendizaje NC post-reconciliacion ─────────────────
        if not df_comp.empty and 'DOC_AUXILIAR' in df_comp.columns:
            _nc_matches = df_comp[
                df_comp['DOC_AUXILIAR'].str.startswith('NC-', na=False)
            ]
            for _, _nr in _nc_matches.iterrows():
                registrar_aprendizaje_nc(
                    str(_nr.get('DESCRIPCION',   '') or ''),
                    str(_nr.get('DOC_AUXILIAR',  '') or ''),
                    str(_nr.get('CONCEPTO_AUX',  '') or ''),
                    str(_nr.get('METODO_MATCH',  'MONTO') or 'MONTO'),
                    _nr.get('VALOR_BANCO'), _nr.get('MONTO_AUXILIAR')
                )
        n_exac = (df_comp['ESTADO'] == '✅ COINCIDE EXACTO').sum()
        n_apr  = (df_comp['ESTADO'] == '🔶 COINCIDE APROX.').sum()
        n_sbco = (df_comp['ESTADO'] == '❌ SOLO EN BANCO').sum()
        n_rec  = (df_comp['ESTADO'] == '🔄 RECHAZO — CONFIRMAR').sum()
        n_agr  = df_comp['ESTADO'].str.startswith('🔵 AGRUPADO', na=False).sum()
        n_saux = len(df_solo_aux)
        pct_conc = (n_exac + n_apr) / max(n_tot, 1) * 100
        exactas = df_comp[df_comp['ESTADO'] == '✅ COINCIDE EXACTO']
        aprox   = df_comp[df_comp['ESTADO'] == '🔶 COINCIDE APROX.']
        s_banco = df_comp[df_comp['ESTADO'] == '❌ SOLO EN BANCO']
    else:
        df_comp = pd.DataFrame()
        df_solo_aux = pd.DataFrame()
        n_tot = n_exac = n_apr = n_sbco = n_saux = n_rec = n_agr = pct_conc = 0
        exactas = aprox = s_banco = rechazos = agrupados = pd.DataFrame()
    # ── Guardar análisis en historial ────────────────────────────────────────
    try:
        _per_det = _extraer_periodo(banco_file.name) if banco_file else None
        _per_str = (f"{_MESES_ES[_per_det[1]-1]} {_per_det[0]}" if _per_det else "")
        guardar_historial({
            "fecha_hora"      : datetime.now().strftime("%Y-%m-%d %H:%M"),
            "archivo_banco"   : banco_file.name if banco_file else "",
            "archivo_auxiliar": aux_file.name   if aux_file   else "",
            "periodo"         : _per_str,
            "n_banco"         : len(df_banco),
            "n_aux"           : len(df_aux),
            "n_exactas"       : int(n_exac),
            "n_aprox"         : int(n_apr),
            "n_solo_banco"    : int(n_sbco),
            "n_solo_aux"      : int(n_saux),
            "tasa"            : float(pct_conc),
            "saldo_banco"     : float(sac  or 0),
            "saldo_aux"       : float(sf_a or 0),
            "diferencia_neta" : float((sac or 0) - (sf_a or 0)),
            "excel_path"      : "",
        })
    except Exception:
        pass


    # ── Pestañas ──────────────────────────────────────────────────────────
    tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs([
        "📊 Diagnóstico", "🏦 Extracto Banco", "📋 Auxiliar Contable",
        "🔗 Comparación", "📝 Diferencias", "⚖️ Conciliación Formal",
        "📈 Visualizaciones", "💾 Exportar Excel"
    ])

    with tab1:
        st.markdown("<div class='section-title'>📊 Diagnóstico de Archivos</div>", unsafe_allow_html=True)
        p_banco, cal_banco, adv_banco, fmt_banco, conf_banco = leg_banco
        p_aux,   cal_aux,   adv_aux,   fmt_aux,   conf_aux   = leg_aux

        c1, c2 = st.columns(2)
        with c1:
            st.metric("🏦 Legibilidad Extracto Banco", f"{p_banco:.1f}%", cal_banco)
            badge_b = "badge-verde" if p_banco >= 90 else ("badge-naranja" if p_banco >= 70 else "badge-rojo")
            st.markdown(f"""
            <div class='callout-info'>
              <b>Formato detectado:</b> {fmt_banco}<br>
              <b>Confianza:</b> <span class='{badge_b}'>{conf_banco}%</span><br>
              <b>Movimientos leídos:</b> {len(df_banco)} registros
            </div>""", unsafe_allow_html=True)
            for a in adv_banco:
                st.warning(a)
        with c2:
            st.metric("📋 Legibilidad Auxiliar Contable", f"{p_aux:.1f}%", cal_aux)
            badge_a = "badge-verde" if p_aux >= 90 else ("badge-naranja" if p_aux >= 70 else "badge-rojo")
            st.markdown(f"""
            <div class='callout-info'>
              <b>Formato detectado:</b> {fmt_aux}<br>
              <b>Confianza:</b> <span class='{badge_a}'>{conf_aux}%</span><br>
              <b>Asientos leídos:</b> {len(df_aux)} registros
            </div>""", unsafe_allow_html=True)
            for a in adv_aux:
                st.warning(a)

        st.markdown("<div class='section-title'>📝 Evaluación de Calidad</div>", unsafe_allow_html=True)
        if p_banco >= 95 and p_aux >= 95:
            st.markdown("""
            <div class='callout-success'>
              <b>✅ Ambos archivos completamente legibles.</b><br>
              Los datos fueron extraídos con alta fidelidad. El análisis de conciliación tiene máxima confiabilidad.
            </div>""", unsafe_allow_html=True)
        elif p_banco >= 80 and p_aux >= 80:
            st.markdown("""
            <div class='callout-warning'>
              <b>⚠️ Legibilidad aceptable con observaciones.</b><br>
              Revise las advertencias anteriores. Algunos campos pueden haber perdido precisión.
            </div>""", unsafe_allow_html=True)
        else:
            st.markdown("""
            <div class='callout-danger'>
              <b>🔴 Legibilidad baja — resultados poco confiables.</b><br>
              Si el PDF está escaneado, active OCR o exporte el archivo a CSV/Excel desde el sistema fuente.
            </div>""", unsafe_allow_html=True)

        st.markdown("<div class='section-title'>ℹ️ Resumen Narrativo</div>", unsafe_allow_html=True)
        fmt_b_lbl = fmt_banco if fmt_banco else "Desconocido"
        fmt_a_lbl = fmt_aux   if fmt_aux   else "Desconocido"
        st.markdown(f"""
        <div class='callout-info'>
          Se cargó el extracto bancario en formato <b>{fmt_b_lbl}</b> (confianza {conf_banco}%)
          con <b>{len(df_banco)} movimientos</b> y el auxiliar contable en formato <b>{fmt_a_lbl}</b>
          (confianza {conf_aux}%) con <b>{len(df_aux)} asientos</b>.
          El sistema detectó los formatos automáticamente sin configuración manual.
        </div>""", unsafe_allow_html=True)

    with tab2:
        st.markdown("<div class='section-title'>🏦 Extracto Bancolombia</div>", unsafe_allow_html=True)
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Saldo Anterior", cop(sa))
        col2.metric("Total Abonos (+)", cop(tab_s))
        col3.metric("Total Cargos (−)", cop(tca_s))
        col4.metric("Saldo Final", cop(sac))

        dif_arit = (sa + tab_s - tca_s) - sac
        if abs(dif_arit) < 1:
            st.markdown(f"""
            <div class='callout-success'>
              <b>✅ El extracto cuadra aritméticamente.</b><br>
              {cop(sa)} + {cop(tab_s)} − {cop(tca_s)} = <b>{cop(sa+tab_s-tca_s)}</b>
              &nbsp;≈&nbsp; Saldo final declarado <b>{cop(sac)}</b>
              &nbsp;·&nbsp; Diferencia: <b>{cop(dif_arit)}</b>
            </div>""", unsafe_allow_html=True)
        else:
            st.markdown(f"""
            <div class='callout-danger'>
              <b>⚠️ El extracto NO cuadra aritméticamente.</b><br>
              {cop(sa)} + {cop(tab_s)} − {cop(tca_s)} = <b>{cop(sa+tab_s-tca_s)}</b>
              &nbsp;vs&nbsp; Saldo declarado <b>{cop(sac)}</b>
              &nbsp;·&nbsp; <b>Diferencia: {cop(dif_arit)}</b>
            </div>""", unsafe_allow_html=True)

        # Análisis de anomalías
        st.markdown("<div class='section-title'>🔍 Análisis de Movimientos</div>", unsafe_allow_html=True)
        n_abonos = (df_banco['TIPO'] == 'ABONO').sum() if 'TIPO' in df_banco.columns else 0
        n_cargos = (df_banco['TIPO'] == 'CARGO').sum() if 'TIPO' in df_banco.columns else 0
        st.markdown(f"""
        <div class='callout-info'>
          Total movimientos: <b>{len(df_banco)}</b>
          &nbsp;·&nbsp; Abonos: <b>{n_abonos}</b>
          &nbsp;·&nbsp; Cargos: <b>{n_cargos}</b>
          &nbsp;·&nbsp; Promedio por movimiento: <b>{cop(df_banco['VALOR'].abs().mean() if not df_banco.empty else 0)}</b>
        </div>""", unsafe_allow_html=True)

        # Top 5 movimientos por valor absoluto
        if not df_banco.empty and 'VALOR' in df_banco.columns:
            top5 = df_banco.nlargest(5, df_banco['VALOR'].abs().name if hasattr(df_banco['VALOR'].abs(), 'name') else 'VALOR')
            try:
                top5 = df_banco.iloc[df_banco['VALOR'].abs().nlargest(5).index]
            except Exception:
                top5 = df_banco.head(5)
            with st.expander("📌 Top 5 movimientos de mayor valor", expanded=False):
                cols_show = [c for c in ['FECHA_RAW','DESCRIPCION','VALOR','SALDO','TIPO'] if c in df_banco.columns]
                st.dataframe(top5[cols_show], use_container_width=True)

        st.markdown("<div class='section-title'>📄 Detalle de Transacciones</div>", unsafe_allow_html=True)
        cols_banco = [c for c in ['FECHA_RAW','DESCRIPCION','VALOR','SALDO','TIPO'] if c in df_banco.columns]
        st.dataframe(df_banco[cols_banco], use_container_width=True, height=400)

    with tab3:
        st.markdown("<div class='section-title'>📋 Auxiliar Contable</div>", unsafe_allow_html=True)
        if not df_aux.empty:
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Saldo Inicial", cop(si_a))
            col2.metric("Total Débitos", cop(td_a))
            col3.metric("Total Créditos", cop(tc_a))
            col4.metric("Saldo Final", cop(sf_a))

            dif_arit_aux = (si_a + td_a - tc_a) - sf_a
            if abs(dif_arit_aux) < 1:
                st.markdown(f"""
                <div class='callout-success'>
                  <b>✅ El auxiliar cuadra aritméticamente.</b><br>
                  {cop(si_a)} + {cop(td_a)} − {cop(tc_a)} = <b>{cop(si_a+td_a-tc_a)}</b>
                  &nbsp;≈&nbsp; Saldo final declarado <b>{cop(sf_a)}</b>
                  &nbsp;·&nbsp; Diferencia: <b>{cop(dif_arit_aux)}</b>
                </div>""", unsafe_allow_html=True)
            else:
                st.markdown(f"""
                <div class='callout-danger'>
                  <b>⚠️ El auxiliar NO cuadra aritméticamente.</b><br>
                  Diferencia de <b>{cop(dif_arit_aux)}</b> entre saldo calculado y declarado.
                  Verifique asientos de apertura o cierres de período.
                </div>""", unsafe_allow_html=True)

            # Desglose por tipo de documento
            deb_df = df_aux[df_aux['DEBITO'].notna()]
            cre_df = df_aux[df_aux['CREDITO'].notna()]
            des_df = df_aux[df_aux['COLUMNA'] == 'DESCONOCIDO'] if 'COLUMNA' in df_aux.columns else pd.DataFrame()

            st.markdown("<div class='section-title'>📊 Desglose por Tipo de Asiento</div>", unsafe_allow_html=True)
            ca, cb, cc = st.columns(3)
            ca.metric("Asientos DÉBITO", len(deb_df), f"{cop(deb_df['DEBITO'].sum())}")
            cb.metric("Asientos CRÉDITO", len(cre_df), f"{cop(cre_df['CREDITO'].sum())}")
            cc.metric("Sin clasificar", len(des_df))

            if len(des_df) > 0:
                st.markdown(f"""
                <div class='callout-warning'>
                  <b>⚠️ {len(des_df)} asientos sin clasificar (columna DESCONOCIDO).</b><br>
                  Estos asientos no pudieron ser identificados como DÉBITO ni CRÉDITO.
                  Pueden afectar el cálculo de la conciliación.
                </div>""", unsafe_allow_html=True)

            # Desglose por tipo de comprobante (primeras 2 letras del documento)
            if 'DOCUMENTO' in df_aux.columns:
                tipo_doc = df_aux['DOCUMENTO'].str[:2].value_counts().head(8)
                if not tipo_doc.empty:
                    with st.expander("📌 Distribución por tipo de comprobante", expanded=False):
                        for prefijo, cnt in tipo_doc.items():
                            st.markdown(f"&nbsp;&nbsp;<span class='badge-azul'>{prefijo}</span> — <b>{cnt}</b> asientos", unsafe_allow_html=True)

            st.markdown("<div class='section-title'>📄 Detalle de Asientos</div>", unsafe_allow_html=True)
            cols_aux = [c for c in ['DOCUMENTO','FECHA_RAW','CONCEPTO','DEBITO','CREDITO','COLUMNA'] if c in df_aux.columns]
            st.dataframe(df_aux[cols_aux], use_container_width=True, height=400)
        else:
            st.markdown("""
            <div class='callout-danger'>
              <b>❌ No se extrajeron asientos del auxiliar contable.</b><br>
              Verifique que el archivo tiene las columnas correctas: Documento, Fecha, Concepto, Débito, Crédito.
              Si es un PDF, active OCR o exporte a CSV desde el sistema contable.
            </div>""", unsafe_allow_html=True)

    with tab4:
        st.markdown("<div class='section-title'>🔗 Comparación Banco ↔ Auxiliar</div>", unsafe_allow_html=True)
        if df_aux.empty:
            st.markdown("<div class='callout-warning'>⚠️ Sin datos del auxiliar para comparar.</div>", unsafe_allow_html=True)
        else:
            c1, c2, c3, c4, c5, c6, c7 = st.columns(7)
            c1.metric("Total Analizados", n_tot)
            c2.metric("✅ Exactos",    n_exac, f"{n_exac/max(n_tot,1)*100:.0f}%")
            c3.metric("🔶 Aprox.",     n_apr,  f"{n_apr/max(n_tot,1)*100:.0f}%")
            c4.metric("🔵 Agrupados",  n_agr,  "N:1 NC" if n_agr else "—")
            c5.metric("🔄 Rechazos",   n_rec,  "Confirmar" if n_rec else "—")
            c6.metric("❌ Solo Banco", n_sbco, f"{n_sbco/max(n_tot,1)*100:.0f}%")
            c7.metric("📋 Solo Aux.",  n_saux)

            ico, lbl, cls = _semaforo_conciliacion(pct_conc)
            st.progress(pct_conc / 100, text=f"{ico} Tasa de conciliación: {pct_conc:.1f}% — {lbl}")
            st.markdown(f"""
            <div class='callout-{"success" if pct_conc>=90 else ("warning" if pct_conc>=75 else "danger")}'>
              <b>{ico} Conciliación {lbl} — {pct_conc:.1f}%</b><br>
              {"Más del 90% de los movimientos tienen correspondencia en el auxiliar. Excelente control contable." if pct_conc>=90
               else ("Entre 75% y 90% de los movimientos conciliados. Hay diferencias puntuales que requieren revisión." if pct_conc>=75
               else "Menos del 75% de los movimientos conciliados. Se requiere revisión detallada del auxiliar contable.")}
            </div>""", unsafe_allow_html=True)
            if n_agr > 0:
                st.markdown(f"""
                <div class='callout-info'>
                  <b>🔵 {int(n_agr)} cargo(s) bancarios agrupados y vinculados a 1 NC del auxiliar (N:1).</b><br>
                  El banco los cobró individualmente; el contador los registró como una sola Nota Contable.
                  Ver detalle en <b>📝 Diferencias</b> → sección <i>Agrupados N:1</i>.
                </div>""", unsafe_allow_html=True)
            if n_rec > 0:
                st.markdown(f"""
                <div class='callout-warning'>
                  <b>🔄 {n_rec} cargo(s) bancario(s) posiblemente vinculados a notas contables de rechazo/devolución.</b><br>
                  El sistema detectó NC- con montos similares (±3%). Revíselos en la pestaña
                  <b>📝 Diferencias</b> → sección <i>Rechazos / Devoluciones — Confirmar</i>.
                </div>""", unsafe_allow_html=True)

            st.markdown("<div class='section-title'>📋 Tabla Completa de Comparación</div>", unsafe_allow_html=True)
            st.dataframe(df_comp, use_container_width=True, height=450)

    with tab5:
        st.markdown("<div class='section-title'>📝 Reporte Detallado de Diferencias</div>", unsafe_allow_html=True)
        if df_aux.empty:
            st.markdown("<div class='callout-warning'>⚠️ Sin datos del auxiliar para comparar.</div>", unsafe_allow_html=True)
        else:
            # ── Calcular desglose abonos / cargos para exactas ───────────────
            if not exactas.empty:
                _ex_ab = exactas[exactas['VALOR_BANCO'] > 0]
                _ex_ca = exactas[exactas['VALOR_BANCO'] < 0]
                _ex_bruto = exactas['VALOR_BANCO'].abs().sum()
                _ex_titulo = (f"✅ Coincidencias Exactas — {len(exactas)} mov. "
                              f"· Bruto: ${_ex_bruto:,.0f} COP")
            else:
                _ex_titulo = "✅ Coincidencias Exactas — 0 movimientos"
                _ex_ab = _ex_ca = pd.DataFrame()
                _ex_bruto = 0

            with st.expander(_ex_titulo, expanded=False):
                if exactas.empty:
                    st.markdown("<div class='callout-warning'>Sin coincidencias exactas.</div>", unsafe_allow_html=True)
                else:
                    st.markdown(f"""
                    <div class='callout-success'>
                      <b>{len(exactas)} movimientos conciliados exactamente</b>
                      por valor y tipo de transacción.<br><br>
                      &nbsp;&nbsp;
                      <span class='badge-verde'>Abonos (+)</span>
                      &nbsp; {len(_ex_ab)} transacciones &nbsp;·&nbsp;
                      Total: <b>${_ex_ab['VALOR_BANCO'].sum():,.0f}</b><br>
                      &nbsp;&nbsp;
                      <span class='badge-rojo'>Cargos (−)</span>
                      &nbsp; {len(_ex_ca)} transacciones &nbsp;·&nbsp;
                      Total: <b>${abs(_ex_ca['VALOR_BANCO'].sum()):,.0f}</b><br><br>
                      Valor bruto total movido: <b>${_ex_bruto:,.0f} COP</b><br>
                      <small style='opacity:.7;'>
                        El valor bruto es la suma de valores absolutos (abonos + cargos).
                        El neto algebraico (abonos − cargos) es
                        ${exactas['VALOR_BANCO'].sum():,.0f} —
                        es normal que sea negativo si los cargos superan los abonos en el período.
                      </small>
                    </div>""", unsafe_allow_html=True)
                    cols_e = [c for c in ['FECHA_BANCO','TIPO_MOV','VALOR_BANCO','DOC_AUXILIAR','MONTO_AUXILIAR'] if c in exactas.columns]
                    st.dataframe(exactas[cols_e].head(100), use_container_width=True)
                    if len(exactas) > 100:
                        st.caption(f"Mostrando primeros 100 de {len(exactas)}. Descargue el Excel para ver todos.")

            with st.expander(f"🔶 Coincidencias Aproximadas — {len(aprox)} movimientos", expanded=False):
                if aprox.empty:
                    st.markdown("<div class='callout-success'>Sin diferencias aproximadas.</div>", unsafe_allow_html=True)
                else:
                    st.markdown(f"""
                    <div class='callout-warning'>
                      <b>{len(aprox)} movimientos con diferencias menores</b> (mismo período, valor cercano).<br>
                      Revise si existen redondeos, diferencias de centavos o asientos de ajuste.
                    </div>""", unsafe_allow_html=True)
                    cols_a = [c for c in ['FECHA_BANCO','TIPO_MOV','VALOR_BANCO','MONTO_AUXILIAR','DIFERENCIA','DOC_AUXILIAR'] if c in aprox.columns]
                    st.dataframe(aprox[cols_a], use_container_width=True)

            # ── SECCIÓN AGRUPADOS N:1 ────────────────────────────────────────
            _bruto_agr = agrupados['VALOR_BANCO'].abs().sum() if not agrupados.empty else 0
            with st.expander(
                f"🔵 Cargos Agrupados N:1 — {int(n_agr)} cargos bancarios → NC únicas · ${_bruto_agr:,.0f} COP",
                expanded=bool(n_agr > 0)
            ):
                if agrupados.empty:
                    st.markdown("<div class='callout-success'>✅ Sin cargos agrupados este período.</div>",
                                unsafe_allow_html=True)
                else:
                    # Resumir por NC vinculada
                    _agr_grupos = agrupados.groupby('DOC_AUXILIAR').agg(
                        N_cargos=('VALOR_BANCO', 'count'),
                        Suma_banco=('VALOR_BANCO', lambda x: x.abs().sum()),
                        NC_concepto=('CONCEPTO_AUX', 'first'),
                        Fecha_NC=('FECHA_AUXILIAR', 'first'),
                    ).reset_index()
                    st.markdown(f"""
                    <div class='callout-info'>
                      <b>🔵 {int(n_agr)} cargo(s) bancarios corresponden a {len(_agr_grupos)} NC del auxiliar.</b><br>
                      El banco cobra individualmente (por cada transacción) y el contador registra
                      una sola NC por el total. El sistema los vinculó automáticamente con tolerancia ±1%.<br><br>
                      <b>¿Requieren acción?</b> No — ya están registrados como nota contable.
                      Solo verifique que la NC del auxiliar esté correctamente fechada.
                    </div>""", unsafe_allow_html=True)
                    st.markdown("**Resumen por NC:**")
                    st.dataframe(_agr_grupos.rename(columns={
                        'DOC_AUXILIAR': 'NC Auxiliar',
                        'N_cargos'    : 'N cargos banco',
                        'Suma_banco'  : 'Total banco ($)',
                        'NC_concepto' : 'Concepto NC',
                        'Fecha_NC'    : 'Fecha NC',
                    }), use_container_width=True)
                    st.markdown("**Detalle de cargos individuales:**")
                    _cols_agr = [c for c in ['FECHA_BANCO','DESCRIPCION','VALOR_BANCO',
                                             'DOC_AUXILIAR','CONCEPTO_AUX','CONFIANZA','ESTADO']
                                 if c in agrupados.columns]
                    st.dataframe(agrupados[_cols_agr], use_container_width=True)

            # ── SECCIÓN RECHAZOS / DEVOLUCIONES ──────────────────────────────
            _bruto_rec = rechazos['VALOR_BANCO'].abs().sum() if not rechazos.empty else 0
            with st.expander(
                f"🔄 Rechazos / Devoluciones — CONFIRMAR — {int(n_rec)} trans. · ${_bruto_rec:,.0f} COP",
                expanded=bool(n_rec > 0)
            ):
                if rechazos.empty:
                    st.markdown("<div class='callout-success'>✅ Sin cargos rechazados detectados este período.</div>", unsafe_allow_html=True)
                else:
                    st.markdown(f"""
                    <div class='callout-warning'>
                      <b>🔄 {n_rec} cargo(s) bancario(s) posiblemente vinculados a NC por rechazo o devolución.</b><br>
                      El sistema los detectó con tolerancia de monto ±3% (para cubrir comisiones bancarias por rechazo).
                      <b>Requieren verificación manual</b> antes de darlos por conciliados.<br><br>
                      <b>Qué hacer:</b> Compare cada fila — si el cargo bancario y la NC corresponden al mismo evento,
                      el contador confirma que ya está registrado. Si no corresponden, cree la NC faltante en el sistema contable.
                    </div>""", unsafe_allow_html=True)
                    _cols_rec = [c for c in ['FECHA_BANCO','DESCRIPCION','VALOR_BANCO',
                                             'DOC_AUXILIAR','FECHA_AUXILIAR','CONCEPTO_AUX',
                                             'MONTO_AUXILIAR','DIFERENCIA','CONFIANZA'] if c in rechazos.columns]
                    st.dataframe(rechazos[_cols_rec], use_container_width=True)
                    st.markdown("""
                    <div class='callout-accion'>
                      <b>📌 ACCIÓN:</b> Si confirma la coincidencia → el cargo queda conciliado (no crear nueva NC).
                      Si NO corresponde → crear NC en SIIGO con el valor exacto del cargo bancario.
                    </div>""", unsafe_allow_html=True)

            # ── SECCIÓN CRÍTICA: Movimientos sin registro contable ──────────
            n_sb = len(s_banco)
            bruto_sb = s_banco['VALOR_BANCO'].abs().sum() if not s_banco.empty else 0
            with st.expander(f"❌ Movimientos Bancarios SIN Registro Contable — {int(n_sb)} trans. · Bruto: ${bruto_sb:,.0f} COP", expanded=bool(n_sb > 0)):
                if s_banco.empty:
                    st.markdown("<div class='callout-success'>✅ Todos los movimientos tienen asiento contable.</div>", unsafe_allow_html=True)
                else:
                    abonos_sb = s_banco[s_banco['VALOR_BANCO'] > 0]
                    cargos_sb = s_banco[s_banco['VALOR_BANCO'] < 0]
                    st.markdown(f"""
                    <div class='callout-danger'>
                      <b>❌ {n_sb} movimientos bancarios no tienen asiento en el auxiliar.</b><br>
                      Abonos sin asiento: <b>{_cop_limpio(abonos_sb['VALOR_BANCO'].sum())}</b> ({len(abonos_sb)} trans.)
                      &nbsp;·&nbsp;
                      Cargos sin asiento: <b>{_cop_limpio(cargos_sb['VALOR_BANCO'].sum())}</b> ({len(cargos_sb)} trans.)<br>
                      <b>Valor bruto no registrado: ${bruto_sb:,.0f} COP</b>
                    </div>""", unsafe_allow_html=True)

                    st.markdown("""
                    <div class='callout-accion'>
                      <b>📌 ¿QUÉ HACER?</b> Para cada fila de abajo: ubique el movimiento en el extracto físico
                      y cree el comprobante contable correspondiente en su sistema (SIIGO / Helisa / World Office).
                    </div>""", unsafe_allow_html=True)

                    st.markdown("<div class='section-title'>Guía de acción por movimiento</div>", unsafe_allow_html=True)
                    for _, row in s_banco.iterrows():
                        st.markdown(_guia_banco_sin_aux(row), unsafe_allow_html=True)

            # ── SECCIÓN CRÍTICA: Asientos sin transacción bancaria ─────────
            n_sa = len(df_solo_aux)
            deb_sa = df_solo_aux['DEBITO'].sum()  if not df_solo_aux.empty else 0
            cre_sa = df_solo_aux['CREDITO'].sum() if not df_solo_aux.empty else 0
            with st.expander(f"📋 Asientos Auxiliar SIN Transacción Bancaria — {int(n_sa)} asientos", expanded=bool(n_sa > 0)):
                if df_solo_aux.empty:
                    st.markdown("<div class='callout-success'>✅ Todos los asientos tienen transacción bancaria.</div>", unsafe_allow_html=True)
                else:
                    st.markdown(f"""
                    <div class='callout-warning'>
                      <b>📋 {n_sa} asientos contables no tienen transacción bancaria correspondiente.</b><br>
                      Débitos sin banco: <b>{_cop_limpio(deb_sa)}</b>
                      &nbsp;·&nbsp;
                      Créditos sin banco: <b>{_cop_limpio(cre_sa)}</b>
                    </div>""", unsafe_allow_html=True)

                    st.markdown("""
                    <div class='callout-accion'>
                      <b>📌 ¿QUÉ HACER?</b> Para cada fila de abajo: busque si el movimiento bancario existe
                      en el extracto (puede ser de otro período). Si no existe, puede ser un asiento de ajuste
                      interno, un pago en efectivo, o requiere anulación.
                    </div>""", unsafe_allow_html=True)

                    st.markdown("<div class='section-title'>Guía de acción por asiento</div>", unsafe_allow_html=True)
                    for _, row in df_solo_aux.iterrows():
                        st.markdown(_guia_aux_sin_banco(row), unsafe_allow_html=True)

    with tab6:
        st.markdown("<div class='section-title'>Conciliacion Bancaria Formal</div>", unsafe_allow_html=True)

        calc_banco = sa + tab_s - tca_s
        dif_b      = calc_banco - sac
        calc_aux   = si_a + td_a - tc_a
        dif_a      = calc_aux - sf_a
        dif_saldos = sac - sf_a

        def _card(ok):
            if ok: return ('rgba(102,187,106,0.13)', '#66bb6a', 'rgba(102,187,106,0.35)')
            return ('rgba(239,83,80,0.13)', '#ef5350', 'rgba(239,83,80,0.35)')

        TL = 'padding:5px 4px;font-size:.88rem;'
        TR = 'padding:5px 4px;font-size:.88rem;text-align:right;'
        TLB = 'padding:7px 4px;font-size:.9rem;font-weight:700;'
        TRB = 'padding:7px 4px;font-size:.9rem;font-weight:800;text-align:right;'

        c1, c2 = st.columns(2)
        with c1:
            bg_b, br_b, sp_b = _card(abs(dif_b) < 1)
            est_b = 'CUADRA' if abs(dif_b) < 1 else 'REVISAR'
            st.markdown(f"""
<div style='background:{bg_b};border-left:4px solid {br_b};border-radius:10px;padding:18px 22px;'>
  <div style='font-size:.95rem;font-weight:800;margin-bottom:12px;'>I. Saldo segun Extracto Bancario</div>
  <table style='width:100%;border-collapse:collapse;'>
    <tr><td style='{TL}'>Saldo anterior</td><td style='{TR}'>{cop(sa)}</td></tr>
    <tr><td style='{TL}'>(+) Total abonos</td><td style='{TR}'>{cop(tab_s)}</td></tr>
    <tr><td style='{TL}'>(-) Total cargos</td><td style='{TR}'>{cop(tca_s)}</td></tr>
    <tr><td colspan='2' style='border-top:1px solid {sp_b};padding:2px 0;'></td></tr>
    <tr><td style='{TLB}'>(=) Saldo calculado</td><td style='{TRB}'>{cop(calc_banco)}</td></tr>
    <tr><td style='{TL}'>(=) Saldo declarado</td><td style='{TR}'>{cop(sac)}</td></tr>
    <tr><td colspan='2' style='border-top:1px solid {sp_b};padding:2px 0;'></td></tr>
    <tr><td style='{TLB}'>Diferencia</td>
        <td style='{TRB}color:{br_b};'>{cop(dif_b)} &nbsp; {est_b}</td></tr>
  </table>
</div>""", unsafe_allow_html=True)

        with c2:
            bg_a, br_a, sp_a = _card(abs(dif_a) < 1)
            est_a = 'CUADRA' if abs(dif_a) < 1 else 'REVISAR'
            st.markdown(f"""
<div style='background:{bg_a};border-left:4px solid {br_a};border-radius:10px;padding:18px 22px;'>
  <div style='font-size:.95rem;font-weight:800;margin-bottom:12px;'>II. Saldo segun Auxiliar Contable</div>
  <table style='width:100%;border-collapse:collapse;'>
    <tr><td style='{TL}'>Saldo inicial</td><td style='{TR}'>{cop(si_a)}</td></tr>
    <tr><td style='{TL}'>(+) Total debitos</td><td style='{TR}'>{cop(td_a)}</td></tr>
    <tr><td style='{TL}'>(-) Total creditos</td><td style='{TR}'>{cop(tc_a)}</td></tr>
    <tr><td colspan='2' style='border-top:1px solid {sp_a};padding:2px 0;'></td></tr>
    <tr><td style='{TLB}'>(=) Saldo calculado</td><td style='{TRB}'>{cop(calc_aux)}</td></tr>
    <tr><td style='{TL}'>(=) Saldo final declarado</td><td style='{TR}'>{cop(sf_a)}</td></tr>
    <tr><td colspan='2' style='border-top:1px solid {sp_a};padding:2px 0;'></td></tr>
    <tr><td style='{TLB}'>Diferencia</td>
        <td style='{TRB}color:{br_a};'>{cop(dif_a)} &nbsp; {est_a}</td></tr>
  </table>
</div>""", unsafe_allow_html=True)

        st.markdown('<br>', unsafe_allow_html=True)
        bg_ds, br_ds, sp_ds = _card(abs(dif_saldos) < 1)
        est_ds = ('Saldos iguales' if abs(dif_saldos) < 1
                  else f'Diferencia de {cop(abs(dif_saldos)).strip()} entre banco y auxiliar')
        st.markdown(f"""
<div style='background:{bg_ds};border-left:4px solid {br_ds};border-radius:10px;padding:18px 22px;'>
  <div style='font-size:.95rem;font-weight:800;margin-bottom:12px;'>III. Diferencia Neta Banco vs Auxiliar</div>
  <table style='width:100%;border-collapse:collapse;'>
    <tr><td style='{TL}'>Saldo banco (final)</td><td style='{TR}'>{cop(sac)}</td></tr>
    <tr><td style='{TL}'>Saldo auxiliar (final)</td><td style='{TR}'>{cop(sf_a)}</td></tr>
    <tr><td colspan='2' style='border-top:1px solid {sp_ds};padding:2px 0;'></td></tr>
    <tr><td style='{TLB}'>DIFERENCIA NETA</td>
        <td style='{TRB}color:{br_ds};letter-spacing:.02em;'>{cop(dif_saldos)}</td></tr>
    <tr><td colspan='2' style='padding:4px 0;'></td></tr>
    <tr><td style='{TL}'>Abonos banco vs Debitos auxiliar</td><td style='{TR}'>{cop(tab_s - td_a)}</td></tr>
    <tr><td style='{TL}'>Cargos banco vs Creditos auxiliar</td><td style='{TR}'>{cop(tca_s - tc_a)}</td></tr>
  </table>
  <div style='margin-top:12px;font-weight:700;font-size:.88rem;color:{br_ds};'>{est_ds}</div>
</div>""", unsafe_allow_html=True)

        if not df_aux.empty:
            val_sin_aux       = s_banco['VALOR_BANCO'].sum()    if not s_banco.empty     else 0
            val_sin_banco_deb = df_solo_aux['DEBITO'].sum()     if not df_solo_aux.empty else 0
            val_sin_banco_cre = df_solo_aux['CREDITO'].sum()    if not df_solo_aux.empty else 0
            ico2, lbl2, _     = _semaforo_conciliacion(pct_conc)
            b2 = 'badge-verde' if pct_conc >= 90 else 'badge-naranja'
            st.markdown('<br>', unsafe_allow_html=True)
            st.markdown(f"""
<div class='callout-info'>
  <b>IV. Composicion de la Diferencia y Conclusion Auditora</b><br>
  Monto banco sin registro auxiliar: <b>{cop(val_sin_aux).strip()}</b><br>
  Monto auxiliar debito sin banco: <b>{cop(val_sin_banco_deb).strip()}</b><br>
  Monto auxiliar credito sin banco: <b>{cop(val_sin_banco_cre).strip()}</b><br>
  Tasa de conciliacion: <span class='{b2}'>{ico2} {pct_conc:.1f}% &mdash; {lbl2}</span><br><br>
  <b>Conclusion:</b> Conciliacion elaborada con base en el extracto Bancolombia y el auxiliar
  contable cuenta 1120.05.01. Tasa {pct_conc:.1f}% &mdash;
  {'los registros estan en orden.' if pct_conc >= 90 else 'existen diferencias que requieren revision antes del cierre.'}
</div>""", unsafe_allow_html=True)

        st.markdown('<br>', unsafe_allow_html=True)
        st.markdown("""
<div style='display:flex;gap:32px;padding:22px 28px;
            background:rgba(66,165,245,0.06);
            border:1px solid rgba(66,165,245,0.2);
            border-radius:12px;'>
  <div style='flex:1;text-align:center;
              border-right:1px solid rgba(66,165,245,0.2);
              padding-right:24px;'>
    <div style='font-weight:800;font-size:.95rem;color:#42a5f5;'>CARLOS ANDRES SILVA VELA</div>
    <div style='font-size:.82rem;opacity:.7;margin-top:3px;'>REPRESENTANTE LEGAL</div>
    <div style='font-size:.78rem;opacity:.5;margin-top:2px;'>C.C. 1061717925</div>
  </div>
  <div style='flex:1;text-align:center;padding-left:24px;'>
    <div style='font-weight:800;font-size:.95rem;color:#42a5f5;'>FERNANDO CUCALON SANCHEZ</div>
    <div style='font-size:.82rem;opacity:.7;margin-top:3px;'>CONTADOR PUBLICO</div>
    <div style='font-size:.78rem;opacity:.5;margin-top:2px;'>T.P. 23049-T</div>
  </div>
</div>""", unsafe_allow_html=True)

    with tab7:
        st.markdown("<div class='section-title'>Visualizaciones</div>", unsafe_allow_html=True)
        plt.rcParams.update({"font.family": "DejaVu Sans", "figure.dpi": 110})
        fig, axes = plt.subplots(2, 3, figsize=(22, 12))
        fig.suptitle("CREDIEXPRESS POPAYAN SAS — Conciliacion Bancaria",
                     fontsize=14, fontweight="bold", y=1.01)

        # G1 — Evolucion del saldo
        ax1 = axes[0, 0]
        df_s = df_banco[df_banco["SALDO"].notna()].copy()
        if not df_s.empty:
            ax1.plot(range(len(df_s)), df_s["SALDO"]/1e6, color="#1565C0", lw=1.2)
            ax1.fill_between(range(len(df_s)), df_s["SALDO"]/1e6, alpha=0.12, color="#1565C0")
        ax1.set_title("Evolucion del Saldo Bancario", fontweight="bold")
        ax1.set_ylabel("Millones COP")
        ax1.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"${x:.0f}M"))
        ax1.grid(True, alpha=0.3)

        # G2 — Pastel estado conciliacion
        ax2 = axes[0, 1]
        if n_tot > 0:
            cont = df_comp["ESTADO"].value_counts()
            color_map = {
                "COINCIDE EXACTO": "#4CAF50",
                "COINCIDE APROX.": "#FFC107",
                "SOLO EN BANCO":   "#F44336",
            }
            cs  = [next((v for k, v in color_map.items() if k in e), "#9E9E9E") for e in cont.index]
            lbl = [e + " (" + str(v) + ")" for e, v in zip(cont.index, cont.values)]
            ax2.pie(cont.values, labels=lbl, colors=cs, autopct="%1.0f%%", startangle=90,
                    textprops={"fontsize": 8})
        ax2.set_title("Estado Conciliacion", fontweight="bold")

        # G3 — Barras Banco vs Auxiliar
        ax3 = axes[0, 2]
        cats  = ["Entradas\nBanco", "Debitos\nAuxiliar", "Salidas\nBanco", "Creditos\nAuxiliar"]
        vals3 = [tab_s/1e6, td_a/1e6, tca_s/1e6, tc_a/1e6]
        cols3 = ["#2196F3", "#4CAF50", "#F44336", "#FF9800"]
        bars3 = ax3.bar(cats, vals3, color=cols3, alpha=0.85)
        ax3.set_title("Totales: Banco vs Auxiliar", fontweight="bold")
        ax3.set_ylabel("Millones COP")
        ax3.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"${x:.0f}M"))
        for b, v in zip(bars3, vals3):
            ax3.text(b.get_x() + b.get_width()/2, b.get_height() + 1,
                     f"${v:.0f}M", ha="center", fontsize=8, fontweight="bold")
        ax3.grid(True, axis="y", alpha=0.3)

        # G4 — Movimientos por dia
        ax4 = axes[1, 0]
        if not df_banco.empty:
            df_banco["DIA"] = df_banco["FECHA_RAW"].apply(
                lambda x: int(str(x).split("/")[0]) if "/" in str(x) else 0)
            por_dia = df_banco.groupby(["DIA", "TIPO"])["VALOR"].sum().unstack(fill_value=0)
            if "ABONO" in por_dia.columns:
                ax4.bar(por_dia.index, por_dia["ABONO"]/1e6,
                        label="Abonos (+)", color="#2E7D32", alpha=0.8)
            if "CARGO" in por_dia.columns:
                ax4.bar(por_dia.index, por_dia["CARGO"].abs()/1e6,
                        label="Cargos (-)", color="#C62828", alpha=0.7)
        ax4.set_title("Movimientos por Dia", fontweight="bold")
        ax4.set_xlabel("Dia del Mes")
        ax4.set_ylabel("Millones COP")
        ax4.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"${x:.0f}M"))
        ax4.legend(fontsize=8)
        ax4.grid(True, axis="y", alpha=0.3)

        # G5 — Asientos por tipo de comprobante
        ax5 = axes[1, 1]
        if not df_aux.empty:
            tipo_cnt = df_aux["DOCUMENTO"].str[:2].value_counts()
            cols5 = ["#4CAF50", "#2196F3", "#FF9800", "#9C27B0", "#F44336", "#00BCD4"][:len(tipo_cnt)]
            bars5 = ax5.bar(tipo_cnt.index, tipo_cnt.values, color=cols5)
            ax5.set_title("Asientos por Tipo (Auxiliar)", fontweight="bold")
            ax5.set_ylabel("N asientos")
            for b, v in zip(bars5, tipo_cnt.values):
                ax5.text(b.get_x() + b.get_width()/2, b.get_height() + 1,
                         str(v), ha="center", fontsize=9, fontweight="bold")
            ax5.grid(True, axis="y", alpha=0.3)
        else:
            ax5.text(0.5, 0.5, "Sin datos auxiliar", ha="center", va="center", fontsize=12)
            ax5.set_title("Asientos por Tipo (Auxiliar)", fontweight="bold")

        # G6 — Valor por estado conciliacion
        ax6 = axes[1, 2]
        if n_tot > 0:
            ve = exactas["VALOR_BANCO"].abs().sum()/1e6 if not exactas.empty else 0
            va = aprox["VALOR_BANCO"].abs().sum()/1e6   if not aprox.empty   else 0
            vs = s_banco["VALOR_BANCO"].abs().sum()/1e6 if not s_banco.empty else 0
            vx = (df_solo_aux["DEBITO"].fillna(0).sum() + df_solo_aux["CREDITO"].fillna(0).sum())/1e6 if not df_solo_aux.empty else 0
            lbl6 = ["Exacto", "Aprox.", "Solo\nbanco", "Solo\nauxiliar"]
            val6 = [ve, va, vs, vx]
            col6 = ["#4CAF50", "#FFC107", "#F44336", "#2196F3"]
            bars6 = ax6.bar(lbl6, val6, color=col6, alpha=0.85)
            ax6.set_title("Valor por Estado Conciliacion", fontweight="bold")
            ax6.set_ylabel("Millones COP")
            ax6.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"${x:.0f}M"))
            for b, v in zip(bars6, val6):
                ax6.text(b.get_x() + b.get_width()/2, b.get_height() + 0.5,
                         f"${v:.1f}M", ha="center", fontsize=8, fontweight="bold")
            ax6.grid(True, axis="y", alpha=0.3)

        plt.tight_layout()
        st.pyplot(fig)

        # Comentario textual debajo de los graficos
        st.markdown("<div class='section-title'>Interpretacion de los Graficos</div>", unsafe_allow_html=True)
        ico_v, lbl_v, _ = _semaforo_conciliacion(pct_conc)
        st.markdown(f"""
        <div class='callout-info'>
          <b>G1 — Evolucion del Saldo:</b> Muestra como vario el saldo bancario durante el periodo.
          Una linea estable indica flujo predecible; caidas o picos bruscos requieren revision.<br><br>
          <b>G2 — Estado de Conciliacion:</b> {ico_v} <b>{pct_conc:.1f}%</b> de los movimientos estan conciliados ({lbl_v}).
          El {100-pct_conc:.1f}% restante esta pendiente de revisionar.<br><br>
          <b>G3 — Banco vs Auxiliar:</b> Compara los totales de entradas/salidas entre el extracto y el auxiliar.
          Barras parejas indican registros completos; barras desiguales apuntan a diferencias.<br><br>
          <b>G4 — Movimientos por Dia:</b> Identifica dias de mayor actividad bancaria.
          Picos de debitos o creditos en fechas especificas pueden indicar pagos masivos o recaudos.<br><br>
          <b>G5 — Tipo de Comprobante:</b> Distribucion de asientos por prefijo de documento (NC=notas credito, CE=comprobantes egreso, etc.).<br><br>
          <b>G6 — Valor por Estado:</b> Muestra el valor monetario en cada categoria de conciliacion.
          El mayor valor debe estar en "Exacto"; valores altos en "Solo banco" o "Solo auxiliar" requieren atencion.
        </div>""", unsafe_allow_html=True)

    with tab8:
        st.markdown("<div class='section-title'>Exportar a Excel</div>", unsafe_allow_html=True)
        st.markdown("""
        <div class='callout-info'>
          El archivo Excel contiene <b>8 hojas</b> con toda la informacion del analisis:
          Comparacion completa, Coincidencias exactas, Aproximadas, Solo en banco, Solo en auxiliar,
          Datos completos del extracto, Datos completos del auxiliar y Resumen ejecutivo.
        </div>""", unsafe_allow_html=True)

        nombre_salida = "CREDIEXPRESS_Conciliacion.xlsx"

        FILL_VERDE    = PatternFill("solid", fgColor="C8F7C5")
        FILL_AMARILLO = PatternFill("solid", fgColor="FFF3CD")
        FILL_ROJO     = PatternFill("solid", fgColor="F7C5C5")
        FILL_AZUL     = PatternFill("solid", fgColor="D0E8FF")
        FILL_NARANJA  = PatternFill("solid", fgColor="FFE0B2")   # rechazos pendientes
        FILL_CELESTE  = PatternFill("solid", fgColor="B3E5FC")   # agrupados N:1
        FILL_HEADER   = PatternFill("solid", fgColor="1565C0")
        FONT_HEADER   = Font(bold=True, color="FFFFFF", size=10)

        def estilizar_hoja(ws):
            for cell in ws[1]:
                cell.fill = FILL_HEADER
                cell.font = FONT_HEADER
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for col in ws.columns:
                mx = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 3, 55)

        def colorear_por_estado(ws, col_estado_idx):
            for row in ws.iter_rows(min_row=2):
                val = str(row[col_estado_idx - 1].value or "")
                fill = (FILL_VERDE    if "COINCIDE EXACTO" in val else
                        FILL_AMARILLO if "COINCIDE APROX"  in val else
                        FILL_CELESTE  if "AGRUPADO"        in val else
                        FILL_NARANJA  if "RECHAZO"         in val else
                        FILL_ROJO     if "SOLO EN BANCO"   in val else
                        FILL_AZUL     if "SOLO EN AUXILIAR" in val else None)
                if fill:
                    for cell in row:
                        cell.fill = fill

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            if not df_aux.empty:
                h1 = df_comp[["N","FECHA_BANCO","TIPO_MOV","DESCRIPCION","VALOR_BANCO",
                               "DOC_AUXILIAR","FECHA_AUXILIAR","CONCEPTO_AUX","MONTO_AUXILIAR",
                               "DIFERENCIA","ESTADO"]].copy()
                h1.columns = ["N","Fecha_Banco","Tipo","Descripcion_Banco","Valor_Banco",
                              "Doc_Auxiliar","Fecha_Auxiliar","Concepto_Auxiliar","Monto_Auxiliar",
                              "Diferencia","Estado"]
            else:
                h1 = pd.DataFrame({"Info": ["Sin comparacion"]})
            h1.to_excel(writer, sheet_name="1_Comparacion_Completa", index=False)

            for estado, nombre in [
                ("COINCIDE EXACTO", "2_Coincidencias_Exactas"),
                ("COINCIDE APROX.", "3_Coincidencias_Aprox"),
                ("AGRUPADO",       "4_Agrupados_N1"),
                ("RECHAZO",        "5_Rechazos_Confirmar"),
                ("SOLO EN BANCO",  "6_Solo_Banco_Sin_Auxiliar"),
            ]:
                sub = df_comp[df_comp["ESTADO"].str.contains(estado, na=False)].copy() if not df_aux.empty else pd.DataFrame()
                if sub.empty: sub = pd.DataFrame({"Info": ["Sin registros"]})
                sub.to_excel(writer, sheet_name=nombre, index=False)

            if not df_solo_aux.empty:
                df_solo_aux.to_excel(writer, sheet_name="7_Solo_Auxiliar_Sin_Banco", index=False)
            else:
                pd.DataFrame({"Info": ["Todos los asientos tienen movimiento bancario"]}).to_excel(
                    writer, sheet_name="7_Solo_Auxiliar_Sin_Banco", index=False)

            df_banco.to_excel(writer, sheet_name="7_Extracto_Banco_Completo", index=True)
            df_aux.to_excel(writer, sheet_name="8_Auxiliar_Contable_Completo", index=True)

            resumen_data = {
                "Concepto": [
                    "Archivo banco", "Archivo auxiliar",
                    "Saldo inicial banco", "Saldo final banco",
                    "Total abonos banco", "Total cargos banco",
                    "Saldo inicial auxiliar", "Saldo final auxiliar",
                    "Total debitos auxiliar", "Total creditos auxiliar",
                    "Diferencia saldos finales",
                    "Movimientos analizados", "Coincidencias exactas",
                    "Coincidencias aprox.", "Agrupados N:1",
                    "Rechazos confirmar", "Solo en banco", "Solo en auxiliar",
                    "Tasa de conciliacion %",
                ],
                "Valor": [
                    banco_file.name, aux_file.name,
                    sa, sac, tab_s, tca_s,
                    si_a, sf_a, td_a, tc_a,
                    sac - sf_a,
                    n_tot, n_exac, n_apr, int(n_agr), int(n_rec), n_sbco, n_saux,
                    round(pct_conc, 1),
                ]
            }
            pd.DataFrame(resumen_data).to_excel(writer, sheet_name="9_Resumen_Conciliacion", index=False)

            wb = writer.book
            for sname in wb.sheetnames:
                ws = wb[sname]
                estilizar_hoja(ws)
            if "1_Comparacion_Completa" in wb.sheetnames and not df_aux.empty:
                ws1 = wb["1_Comparacion_Completa"]
                colorear_por_estado(ws1, 11)

        output.seek(0)
        # ── Auto-guardar Excel localmente (solo offline) ─────────────────────
        if OFFLINE_MODE:
            _excel_local = _auto_guardar_excel(output.getvalue(), nombre_salida)
            if _excel_local:
                st.markdown(f"""
<div class='callout-success' style='margin-bottom:10px;font-size:.85rem;'>
  💾 Excel guardado automaticamente en:<br>
  <code style='font-size:.78rem;word-break:break-all;'>{_excel_local}</code>
</div>""", unsafe_allow_html=True)


        col_dl, col_info = st.columns([1, 2])
        with col_dl:
            st.download_button(
                label="Descargar Excel Premium",
                data=output,
                file_name=nombre_salida,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        with col_info:
            st.markdown(f"""
            <div class='callout-success'>
              Archivo listo: <b>{nombre_salida}</b><br>
              Movimientos banco: <b>{len(df_banco)}</b> &nbsp;|&nbsp;
              Asientos auxiliar: <b>{len(df_aux)}</b> &nbsp;|&nbsp;
              Conciliacion: <b>{pct_conc:.1f}%</b>
            </div>""", unsafe_allow_html=True)

    st.markdown("""
    <div class='callout-success' style='margin-top:20px;'>
      <b>Analisis completado exitosamente.</b>
      Revise las pestanas para el detalle completo. Descargue el Excel para el archivo oficial.
    </div>""", unsafe_allow_html=True)
