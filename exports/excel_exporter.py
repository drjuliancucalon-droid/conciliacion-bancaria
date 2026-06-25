"""
Exportador a Excel con formato profesional
CREDIEXPRESS POPAYÁN SAS — Conciliación Bancaria
"""

import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import pandas as pd


# ════════════════════════════════════════════════════════════════════════════════
# ESTILOS
# ════════════════════════════════════════════════════════════════════════════════

HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(name='Calibri', bold=True, size=14, color='1F4E79')
SUBTITLE_FONT = Font(name='Calibri', bold=True, size=11, color='2E75B6')
NORMAL_FONT = Font(name='Calibri', size=10)
BOLD_FONT = Font(name='Calibri', bold=True, size=10)
MONEY_FMT = '#,##0.00'
PCT_FMT = '0.0%'

# Colores por estado
FILL_EXACTA = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Verde claro
FILL_APROX = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')   # Amarillo claro
FILL_NC = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')      # Azul claro
FILL_AGRUP = PatternFill(start_color='F2DCDB', end_color='F2DCDB', fill_type='solid')   # Rojo claro
FILL_RECH = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')    # Verde menta
FILL_SOLO_BANCO = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')  # Rosa claro
FILL_SOLO_AUX = PatternFill(start_color='E8EAF6', end_color='E8EAF6', fill_type='solid')     # Índigo claro

THIN_BORDER = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7'),
)

CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')


def _apply_header_style(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def _apply_data_style(ws, start_row, end_row, max_col, money_cols=None, pct_cols=None):
    money_cols = money_cols or []
    pct_cols = pct_cols or []
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = RIGHT_ALIGN if c in money_cols else LEFT_ALIGN
            if c in money_cols:
                cell.number_format = MONEY_FMT
            if c in pct_cols:
                cell.number_format = PCT_FMT


def _auto_width(ws, max_col, min_width=10, max_width=50):
    for col in range(1, max_col + 1):
        max_len = min_width
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[get_column_letter(col)].width = max_len


# ════════════════════════════════════════════════════════════════════════════════
# EXPORTADOR PRINCIPAL
# ════════════════════════════════════════════════════════════════════════════════

def generar_excel_conciliacion(
    matches_df, solo_banco_df, solo_aux_df, stats,
    banco_meta, aux_meta, banco_fmt, aux_fmt,
    periodo, archivo_banco, archivo_auxiliar
):
    """
    Genera el archivo Excel completo de conciliación.
    Retorna bytes del archivo.
    """
    wb = Workbook()
    
    # ═══════════════════════════════════════════════════════════════════════════════
    # HOJA 1: RESUMEN EJECUTIVO
    # ═══════════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = 'Resumen'
    
    # Título
    ws1.merge_cells('A1:H1')
    ws1['A1'] = 'CONCILIACIÓN BANCARIA — CREDIEXPRESS POPAYÁN SAS'
    ws1['A1'].font = TITLE_FONT
    ws1['A1'].alignment = Alignment(horizontal='center')
    
    ws1.merge_cells('A2:H2')
    ws1['A2'] = f'Período: {periodo}  |  Banco: {archivo_banco}  |  Auxiliar: {archivo_auxiliar}'
    ws1['A2'].font = SUBTITLE_FONT
    ws1['A2'].alignment = Alignment(horizontal='center')
    
    # KPIs
    row = 4
    kpis = [
        ('Movimientos Banco', stats['n_banco']),
        ('Asientos Auxiliar', stats['n_aux']),
        ('Matches Exactos', stats['n_exactas']),
        ('Matches Aproximados', stats['n_aprox']),
        ('Solo en Banco', stats['n_solo_banco']),
        ('Solo en Auxiliar', stats['n_solo_aux']),
        ('Tasa de Conciliación', f"{stats['tasa']:.1f}%"),
        ('Saldo Banco', stats['saldo_banco']),
        ('Saldo Auxiliar', stats['saldo_aux']),
        ('Diferencia Neta', stats['diferencia_neta']),
    ]
    
    for label, val in kpis:
        ws1.cell(row=row, column=1, value=label).font = BOLD_FONT
        ws1.cell(row=row, column=1).border = THIN_BORDER
        c = ws1.cell(row=row, column=2, value=val)
        c.border = THIN_BORDER
        if isinstance(val, (int, float)) and label not in ('Tasa de Conciliación',):
            c.number_format = MONEY_FMT
        row += 1
    
    # Semáforo
    from utils.formatters import semaforo_conciliacion
    sem = semaforo_conciliacion(stats['tasa'], stats['n_solo_banco'], stats['n_solo_aux'])
    ws1.cell(row=row, column=1, value='Estado').font = BOLD_FONT
    ws1.cell(row=row, column=2, value=f'{sem} {"EXCELENTE" if sem=="🟢" else "ACEPTABLE" if sem=="🟡" else "REVISAR"}').font = Font(size=12, bold=True)
    
    _auto_width(ws1, 8)
    
    # ═══════════════════════════════════════════════════════════════════════════════
    # HOJA 2: MATCHES EXACTOS
    # ═══════════════════════════════════════════════════════════════════════════════
    if not matches_df.empty:
        exactas = matches_df[matches_df['tipo'] == 'EXACTA'].copy()
        if not exactas.empty:
            ws2 = wb.create_sheet('Exactas')
            cols = ['banco_idx', 'aux_idx', 'valor_banco', 'valor_aux', 'diff',
                    'concepto_banco', 'concepto_aux', 'documento_aux']
            headers = ['Idx Banco', 'Idx Aux', 'Valor Banco', 'Valor Aux', 'Diferencia',
                       'Concepto Banco', 'Concepto Aux', 'Documento']
            
            for c, h in enumerate(headers, 1):
                ws2.cell(row=1, column=c, value=h)
            _apply_header_style(ws2, 1, len(headers))
            
            for i, (_, m) in enumerate(exactas.iterrows(), 2):
                for c, col in enumerate(cols, 1):
                    ws2.cell(row=i, column=c, value=m.get(col, ''))
            
            money_cols = [3, 4, 5]
            _apply_data_style(ws2, 2, len(exactas) + 1, len(headers), money_cols)
            for row in ws2.iter_rows(min_row=2, max_row=len(exactas)+1, min_col=1, max_col=len(headers)):
                for cell in row:
                    cell.fill = FILL_EXACTA
            _auto_width(ws2, len(headers))
    
    # ═══════════════════════════════════════════════════════════════════════════════
    # HOJA 3: MATCHES APROXIMADOS
    # ═══════════════════════════════════════════════════════════════════════════════
    if not matches_df.empty:
        aprox = matches_df[matches_df['tipo'].isin(['APROX', 'NC_CATALOGO', 'AGRUPADO', 'RECHAZO'])].copy()
        if not aprox.empty:
            ws3 = wb.create_sheet('Aproximados')
            cols = ['banco_idx', 'aux_idx', 'tipo', 'valor_banco', 'valor_aux', 'diff',
                    'concepto_banco', 'concepto_aux', 'documento_aux']
            headers = ['Idx Banco', 'Idx Aux', 'Tipo', 'Valor Banco', 'Valor Aux', 'Diferencia',
                       'Concepto Banco', 'Concepto Aux', 'Documento']
            
            for c, h in enumerate(headers, 1):
                ws3.cell(row=1, column=c, value=h)
            _apply_header_style(ws3, 1, len(headers))
            
            fill_map = {
                'APROX': FILL_APROX,
                'NC_CATALOGO': FILL_NC,
                'AGRUPADO': FILL_AGRUP,
                'RECHAZO': FILL_RECH,
            }
            
            for i, (_, m) in enumerate(aprox.iterrows(), 2):
                for c, col in enumerate(cols, 1):
                    ws3.cell(row=i, column=c, value=m.get(col, ''))
                # Colorear por tipo
                tipo = m.get('tipo', '')
                fill = fill_map.get(tipo, FILL_APROX)
                for c in range(1, len(headers) + 1):
                    ws3.cell(row=i, column=c).fill = fill
            
            money_cols = [4, 5, 6]
            _apply_data_style(ws3, 2, len(aprox) + 1, len(headers), money_cols)
            _auto_width(ws3, len(headers))
    
    # ═══════════════════════════════════════════════════════════════════════════════
    # HOJA 4: SOLO EN BANCO
    # ═══════════════════════════════════════════════════════════════════════════════
    if not solo_banco_df.empty:
        ws4 = wb.create_sheet('Solo_Banco')
        cols = ['FECHA_RAW', 'DESCRIPCION', 'VALOR', 'SALDO', 'TIPO', 'PAGINA']
        headers = ['Fecha', 'Descripción', 'Valor', 'Saldo', 'Tipo', 'Página']
        
        for c, h in enumerate(headers, 1):
            ws4.cell(row=1, column=c, value=h)
        _apply_header_style(ws4, 1, len(headers))
        
        for i, (_, r) in enumerate(solo_banco_df.iterrows(), 2):
            for c, col in enumerate(cols, 1):
                val = r.get(col, '')
                if col == 'FECHA_RAW':
                    val = str(val)
                ws4.cell(row=i, column=c, value=val)
        
        money_cols = [3, 4]
        _apply_data_style(ws4, 2, len(solo_banco_df) + 1, len(headers), money_cols)
        for row in ws4.iter_rows(min_row=2, max_row=len(solo_banco_df)+1, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.fill = FILL_SOLO_BANCO
        _auto_width(ws4, len(headers))
    
    # ═══════════════════════════════════════════════════════════════════════════════
    # HOJA 5: SOLO EN AUXILIAR
    # ═══════════════════════════════════════════════════════════════════════════════
    if not solo_aux_df.empty:
        ws5 = wb.create_sheet('Solo_Auxiliar')
        cols = ['DOCUMENTO', 'FECHA_RAW', 'CONCEPTO', 'DEBITO', 'CREDITO', 'COLUMNA', 'VALOR_NETO']
        headers = ['Documento', 'Fecha', 'Concepto', 'Débito', 'Crédito', 'Columna', 'Valor Neto']
        
        for c, h in enumerate(headers, 1):
            ws5.cell(row=1, column=c, value=h)
        _apply_header_style(ws5, 1, len(headers))
        
        for i, (_, r) in enumerate(solo_aux_df.iterrows(), 2):
            for c, col in enumerate(cols, 1):
                val = r.get(col, '')
                if col == 'FECHA_RAW':
                    val = str(val)
                ws5.cell(row=i, column=c, value=val)
        
        money_cols = [4, 5, 7]
        _apply_data_style(ws5, 2, len(solo_aux_df) + 1, len(headers), money_cols)
        for row in ws5.iter_rows(min_row=2, max_row=len(solo_aux_df)+1, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.fill = FILL_SOLO_AUX
        _auto_width(ws5, len(headers))
    
    # ═══════════════════════════════════════════════════════════════════════════════
    # HOJA 6: DETALLE BANCO COMPLETO
    # ═══════════════════════════════════════════════════════════════════════════════
    # (Se requiere el df_banco original - se pasa por parámetro en la llamada real)
    
    # ═══════════════════════════════════════════════════════════════════════════════
    # HOJA 7: DETALLE AUXILIAR COMPLETO
    # ═══════════════════════════════════════════════════════════════════════════════
    
    # ══════════════════════════════════════════════════════════════════════════════
    # HOJA 8: METADATOS Y FORMATOS
    # ═══════════════════════════════════════════════════════════════════════════════
    ws8 = wb.create_sheet('Metadatos')
    ws8.cell(row=1, column=1, value='METADATOS DE LA CONCILIACIÓN').font = TITLE_FONT
    
    meta_info = [
        ('Período', periodo),
        ('Archivo Banco', archivo_banco),
        ('Formato Banco', banco_fmt),
        ('Archivo Auxiliar', archivo_auxiliar),
        ('Formato Auxiliar', aux_fmt),
        ('Fecha Proceso', pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')),
    ]
    
    if banco_meta:
        for k, v in banco_meta.items():
            meta_info.append((f'Banco: {k}', v))
    if aux_meta:
        for k, v in aux_meta.items():
            meta_info.append((f'Aux: {k}', v))
    
    for i, (label, val) in enumerate(meta_info, 3):
        ws8.cell(row=i, column=1, value=label).font = BOLD_FONT
        ws8.cell(row=i, column=1).border = THIN_BORDER
        c = ws8.cell(row=i, column=2, value=val)
        c.border = THIN_BORDER
        if isinstance(val, (int, float)):
            c.number_format = MONEY_FMT
    
    _auto_width(ws8, 2)
    
    # Guardar a bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()